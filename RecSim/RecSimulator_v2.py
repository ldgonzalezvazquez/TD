#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations
import argparse, gzip, json, os, re, shutil, subprocess, sys, tarfile, io, csv, pprint, pickle, sys, itertools
import numpy as _np
import pandas as _pd
import pandas as pd, openpyxl as _xl
from pandas import ExcelWriter
import re
from pathlib import Path
from collections import defaultdict, Counter
from urllib.request import urlretrieve
from Bio import Phylo
from Bio import SeqIO
from typing import Optional, Dict, Any, List, Tuple, Union, Set
from itertools import chain
from Bio.SeqRecord import SeqRecord
import xml.etree.ElementTree as ET
from openpyxl import load_workbook

# ────────────── Utils ────────────── 
_GREEN = "\033[32m"
_RESET = "\033[0m"
EXCEL_PATH = Path("results/recombination_summary.xlsx")

# ──────────────────────────────────────────────────────────────
# Written by Luis Daniel González-Vázquez
# Email: luisdaniel.gonzalez@uvigo.es // luisde365@gmail.com
# Dates: April-Nov 2025
# ──────────────────────────────────────────────────────────────
def print_column_legend() -> None:
    """
    Pretty-prints (to stdout) a full legend for the summary spreadsheets that
    the pipeline generates.  Besides the column-by-column glossary, it also
    reminds the user what each stage of the end-to-end workflow did.
    """
    # ────────────────────────────────────────────────────────────────────
    #  ❷  Column-by-column legend (Excel ‘santa_vs_tools.xlsx’ sheet)
    #      Calculation rules in ⟨…⟩ brackets
    # ────────────────────────────────────────────────────────────────────
    hdr = [
        "tag", "event", "tool", "ext_id",
        "par1_match_n",   "par1_missing_santa", "par1_missing_tool", "par1_match_pct",
        "par2_match_n",   "par2_missing_santa", "par2_missing_tool", "par2_match_pct",
        "rec_match_n",    "rec_missing_santa",  "rec_missing_tool",  "rec_match_pct",
        "nonmatch_sim_prod_p1", "nonmatch_sim_prod_p2",
        "break_overlap_nt", "break_overlap_pct", "break_bp_sizes"
    ]

    _desc = {
        "tag"  : "Run identifier  ➜  <idx>_<nSeq>_<reference>",
        "event": "Santa-Sim event number (ground truth)",
        "tool" : "External method being compared (CFML, Bacter, RDP)",
        "ext_id": "Event identifier **in that tool** "
                  "(node label, group#, or begin–end coordinates)",

        # --- parental matches ------------------------------------------------
        "par1_match_n":
            "# parental sequences that are shared between Santa-Sim Par1 set "
            "and the tool group assigned as Par1  ⟨|P1∩T|⟩",
        "par1_missing_santa":
            "Sequences present in Santa Par1 but absent from the tool "
            "⟨|P1\\T|⟩",
        "par1_missing_tool":
            "Sequences present in the tool group but not in Santa Par1 "
            "⟨|T\\P1|⟩",
        "par1_match_pct":
            "Percentage of Santa Par1 sequences recovered by the tool  "
            "⟨100·|P1∩T| / |P1|⟩",

        # --- Par2 equivalents ------------------------------------------------
        "par2_match_n"          : "Same as above, for Santa Par2",
        "par2_missing_santa"    : "Santa-only Par2 sequences ⟨|P2\\T|⟩",
        "par2_missing_tool"     : "Tool-only Par2 sequences ⟨|T\\P2|⟩",
        "par2_match_pct"        : "100·|P2∩T| / |P2|",

        # --- Recombinants ----------------------------------------------------
        "rec_match_n"           : "# recombinant tips detected by the tool",
        "rec_missing_santa"     : "Recombinants in Santa but not in tool",
        "rec_missing_tool"      : "Tips flagged as rec. by the tool but not Santa",
        "rec_match_pct"         : "100·|REC∩T| / |REC|",

        # --- Similarity products (penalty for ‘wrong’ extra sequences) -------
        "nonmatch_sim_prod_p1":
            "Product of pairwise similarities for sequences that the tool put\n"
            "in Par1 **but Santa did not**  ⟨∏ sim(seq, true Par1)⟩. ",
        "nonmatch_sim_prod_p2":
            "Same for Par2 extras",

        # --- Breakpoint overlap metrics -------------------------------------
        "break_overlap_nt":
            "Raw size of the intersection between Santa’s union-of-fragments\n"
            "and the best-matching tool interval (in nucleotides)",
        "break_overlap_pct":
            "Dice overlap of the **best Santa alternative** vs tool interval\n"
            "⟨100 × (2·|∩| / (|S|+|T|))⟩",
        "break_bp_sizes":
            "Comma-separated sizes of (Santa_breakpoint , tool_breakpoint sizes) that were\n"
            "used when computing the Dice breakpoint overlap.  Handy when you\n"
            "need to eyeball why two events match poorly even with a high pct.",
    }

    print("────────  COLUMN LEGEND (Excel)  ──────────")
    for col in hdr:
        expl = _desc.get(col, "⚠  undocumented column")
        # align the arrow for readability
        print(f"• {col:<24} → {expl}")
    print("──────────────────────────────────────────\n")

def green(msg: str) -> str:
    """Returns the string in green for stdout."""
    return f"{_GREEN}{msg}{_RESET}"

def shell(cmd, cwd=None, env=None, ignore_err=False):
    """
    Launches 'cmd'. If ignore_err=True, does NOT raise exception but
    returns the exit code (0 = OK). Useful for capturing 137.
    """
    try:
        subprocess.run(cmd, check=True, cwd=cwd, env=env)
        return 0
    except subprocess.CalledProcessError as e:
        if ignore_err:
            return e.returncode
        raise

def download(url: str, dest: Path):
    dest.parent.mkdir(parents=True, exist_ok=True)
    if dest.exists():
        return
    print(f"  ↳downloading {url.split('/')[-1]} …")
    urlretrieve(url, dest)

# ───────────────────────────── software ───────────────────────────────
def download_software(soft_dir: Path):
    """
    Downloads and compiles (or extracts) everything needed in ./software/.
    Requires gcc, make, git, tar and Java ≥8 in $PATH.
    """
    soft_dir.mkdir(exist_ok=True)

    # 1) SANTA Luis_MOD
    santa_bin = soft_dir / "santa-sim-Recomb_and_align_Luis_MOD" / "dist" / "santa.jar"
    if not santa_bin.exists():
        shell("ant",cwd=soft_dir/"santa-sim-Recomb_and_align_Luis_MOD")

    # 2) IQ‑TREE 2
    iq_bin = soft_dir / "iqtree-2.2.0-Linux" / "bin" / "iqtree2"
    if not iq_bin.exists():
        tarfile_path = soft_dir / "iqtree-2.2.0-Linux.tar.gz"
        download("https://github.com/iqtree/iqtree2/releases/download/v2.2.0/iqtree-2.2.0-Linux.tar.gz",tarfile_path)
        with tarfile.open(tarfile_path) as tf:
            tf.extractall(path=soft_dir)

    # 3) ClonalFrameML
    cfml_bin = soft_dir / "ClonalFrameML" / "src" / "ClonalFrameML"
    if not cfml_bin.exists():
        shell(["git", "clone", "--depth", "1",
               "https://github.com/xavierdidelot/ClonalFrameML",
               str(soft_dir / "ClonalFrameML")])
        shell(["make"], cwd=soft_dir / "ClonalFrameML" / "src")

    # 4) BEAST 2.6.7 (+ Bacter en lib/launcher.jar)
    beast_dir = soft_dir / "beast2.7.7"
    beast_bin  = beast_dir / "bin" / "beast"
    if not beast_bin.exists():
        tar_path = soft_dir / "BEAST.v2.7.7.Linux.x86.tgz"
        shell(["wget","https://github.com/CompEvol/beast2/releases/download/v2.7.7/BEAST.v2.7.7.Linux.x86.tgz"],cwd=soft_dir)
        with tarfile.open(tar_path) as tf:
            tf.extractall(path=soft_dir)
        shell(["mv","beast","beast2.7.7"],cwd=soft_dir)

    beast_bin  = beast_dir / "bin" / "packagemanager"
    shell([beast_bin,"-add","bacter"],cwd=soft_dir)
    shell([beast_bin,"-add","BEASTLabs"],cwd=soft_dir)
    shell([beast_bin,"-add","MODEL_SELECTION"],cwd=soft_dir)
    

    print("✓Software downloaded and ready in ./software/")
    sys.exit(0)

# ───────────────────────────── SANTA‑SIM XML ──────────────────────────
def create_santa_config_xml(ref_fasta, xml_path,
                            pop_size, mut_rate, recomb_prob,
                            dual_infect_prob, gen_count, sample_size):
    """writes (once) cfg.xml inside each santa/ folder"""
    #if xml_path.exists():
    #    return
    seq = "".join(l.strip() for l in open(ref_fasta) if not l.startswith(">"))
    L   = len(seq)
    xml = f"""<santa xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
       xsi:noNamespaceSchemaLocation="santa.xsd">

    <replicates>1</replicates>

    <simulation>
        <genome>
            <length>{L}</length>
            <sequences>{seq}</sequences>
            <feature><name>GENOME</name><type>nucleotide</type><coordinates>1-{L}</coordinates></feature>
        </genome>

        <population><populationSize>{pop_size}</populationSize><inoculum>all</inoculum></population>
        <populationType>staticPopulation</populationType>
        <fitnessFunction><neutralFitness/></fitnessFunction>

        <mutator><nucleotideMutator><mutationRate>{mut_rate}</mutationRate><transitionBias>1.0</transitionBias></nucleotideMutator></mutator>

        <replicator><recombinantReplicator>
            <dualInfectionProbability>{dual_infect_prob}</dualInfectionProbability>
            <recombinationProbability>{recomb_prob}</recombinationProbability>
        </recombinantReplicator></replicator>

        <epoch><name>epoch1</name><generationCount>{gen_count}</generationCount></epoch>

        <samplingSchedule>
            <sampler><atGeneration>{gen_count}</atGeneration><fileName>final_sample.fasta</fileName>
                <alignment><sampleSize>{sample_size}</sampleSize><format>FASTA</format><label>seq_%s</label></alignment>
            </sampler>
            <sampler><atGeneration>{gen_count}</atGeneration><fileName>final_tree.newick</fileName>
                <tree><sampleSize>{sample_size}</sampleSize><format>NEWICK</format><label>seq_%s</label></tree>
            </sampler>
            <sampler><atFrequency>1</atFrequency><fileName>stats.csv</fileName><statistics/></sampler>
        </samplingSchedule>
    </simulation>
</santa>"""
    xml_path.write_text(xml)


# ─────────────────────── Bacter XML minimal ───────────────────────────

def build_bacter_xml(aln: Path, xml_out: Path,
                     tag: str, chain: int,
                     delta: float, rho: float) -> None:
    """
    Creates an XML file for BEAST-Bacter *identical* to the provided model.
    The only things that change are:
        • tag     → alignment id, injected in dozens of places
        • chain   → MCMC chainLength
        • delta   → initial value of δ
        • rho     → initial value of ρ
        • seq_block → actual sequences from the FASTA
    """
    # ------------------------------------------------------------------
    # 1. Build the sequence block EXACTLY with the format
    #    of the target XML (4 spaces indentation before <sequence …/>)
    # ------------------------------------------------------------------
    seq_lines = []
    with aln.open() as fa:
        it = iter(fa)
        for line in it:
            if line.startswith(">"):
                name = line[1:].strip()
                seq  = next(it).strip()
                seq_lines.append(
                    f'        <sequence id="s_{name}" spec="Sequence" '
                    f'taxon="{name}" totalcount="4" value="{seq}"/>'
                )
    seq_block = "\n".join(seq_lines) or "        <!-- FASTA empty -->"

    # ------------------------------------------------------------------
    # 2. XML template 100% faithful to the example.
    #    DO NOT touch anything except the {{…}} markers!
    # ------------------------------------------------------------------
    template = f"""<?xml version="1.0" encoding="UTF-8" standalone="no"?><beast beautitemplate='Bacter' beautistatus='' namespace="beast.core:beast.evolution.alignment:beast.evolution.tree.coalescent:beast.core.util:beast.evolution.nuc:beast.evolution.operators:beast.evolution.sitemodel:beast.evolution.substitutionmodel:beast.base.evolution.alignment:beast.pkgmgmt:beast.base.core:beast.base.inference:beast.base.evolution.tree.coalescent:beast.pkgmgmt:beast.base.core:beast.base.inference.util:beast.evolution.nuc:beast.base.evolution.operator:beast.base.inference.operator:beast.base.evolution.sitemodel:beast.base.evolution.substitutionmodel:beast.base.evolution.likelihood" required="BEAST.base v2.7.7:bacter v3.0.1" version="2.7">

    <data
id="{tag}"
spec="Alignment"
name="alignment">
{seq_block}     </data>

    <map name="Uniform" >beast.base.inference.distribution.Uniform</map>

    <map name="Exponential" >beast.base.inference.distribution.Exponential</map>

    <map name="LogNormal" >beast.base.inference.distribution.LogNormalDistributionModel</map>

    <map name="Normal" >beast.base.inference.distribution.Normal</map>

    <map name="Beta" >beast.base.inference.distribution.Beta</map>

    <map name="Gamma" >beast.base.inference.distribution.Gamma</map>

    <map name="LaplaceDistribution" >beast.base.inference.distribution.LaplaceDistribution</map>

    <map name="prior" >beast.base.inference.distribution.Prior</map>

    <map name="InverseGamma" >beast.base.inference.distribution.InverseGamma</map>

    <map name="OneOnX" >beast.base.inference.distribution.OneOnX</map>

    <run id="mcmc" spec="MCMC" chainLength="{chain}" storeEvery="10000">
        <state id="state" spec="State" storeEvery="10000">
            <stateNode id="Tree.t:{tag}" spec="bacter.model.SimulatedACG" delta="{delta}" rho="{rho}">
                <populationModel id="ConstantPopulation.0" spec="ConstantPopulation">
                    <parameter id="RealParameter.0" spec="parameter.RealParameter" name="popSize">1.0</parameter>
                </populationModel>
                <locus id="locus.{tag}" spec="bacter.Locus">
                    <alignment idref="{tag}"/>
                </locus>
                <taxonset id="TaxonSet.1" spec="TaxonSet" alignment="@{tag}"/>
            </stateNode>
            <parameter id="rateAC.s:{tag}" spec="parameter.RealParameter" lower="0.0" name="stateNode">1.0</parameter>
            <parameter id="rateAG.s:{tag}" spec="parameter.RealParameter" lower="0.0" name="stateNode">1.0</parameter>
            <parameter id="rateAT.s:{tag}" spec="parameter.RealParameter" lower="0.0" name="stateNode">1.0</parameter>
            <parameter id="rateCG.s:{tag}" spec="parameter.RealParameter" lower="0.0" name="stateNode">1.0</parameter>
            <parameter id="rateGT.s:{tag}" spec="parameter.RealParameter" lower="0.0" name="stateNode">1.0</parameter>
            <parameter id="freqParameter.s:{tag}" spec="parameter.RealParameter" dimension="4" lower="0.0" name="stateNode" upper="1.0">0.25</parameter>
            <parameter id="rho.t:{tag}" spec="parameter.RealParameter" name="stateNode">{rho}</parameter>
            <parameter id="delta.t:{tag}" spec="parameter.RealParameter" name="stateNode">{delta}</parameter>
            <parameter id="popSize.t:{tag}" spec="parameter.RealParameter" name="stateNode">1.0</parameter>
        </state>
        <distribution id="posterior" spec="CompoundDistribution">
            <distribution id="prior" spec="CompoundDistribution">
                <distribution id="CoalescentConstant.t:{tag}" spec="bacter.model.ACGCoalescent" delta="@delta.t:{tag}" rho="@rho.t:{tag}" tree="@Tree.t:{tag}">
                    <populationModel id="popModelConstant.t:{tag}" spec="ConstantPopulation" popSize="@popSize.t:{tag}"/>
                </distribution>
                <prior id="deltaPrior.t:{tag}" name="distribution" x="@delta.t:{tag}">
                    <OneOnX id="OneOnX.1" name="distr"/>
                </prior>
                <prior id="FrequenciesPrior.s:{tag}" name="distribution" x="@freqParameter.s:{tag}">
                    <distr id="Dirichlet.0" spec="distribution.Dirichlet">
                        <parameter id="RealParameter.13" spec="parameter.RealParameter" dimension="4" estimate="false" name="alpha">4.0 4.0 4.0 4.0</parameter>
                    </distr>
                </prior>
                <prior id="RateACPrior.s:{tag}" name="distribution" x="@rateAC.s:{tag}">
                    <Gamma id="Gamma.0" name="distr">
                        <parameter id="RealParameter.1" spec="parameter.RealParameter" estimate="false" name="alpha">0.05</parameter>
                        <parameter id="RealParameter.2" spec="parameter.RealParameter" estimate="false" name="beta">10.0</parameter>
                    </Gamma>
                </prior>
                <prior id="RateAGPrior.s:{tag}" name="distribution" x="@rateAG.s:{tag}">
                    <Gamma id="Gamma.1" name="distr">
                        <parameter id="RealParameter.3" spec="parameter.RealParameter" estimate="false" name="alpha">0.05</parameter>
                        <parameter id="RealParameter.4" spec="parameter.RealParameter" estimate="false" name="beta">20.0</parameter>
                    </Gamma>
                </prior>
                <prior id="RateATPrior.s:{tag}" name="distribution" x="@rateAT.s:{tag}">
                    <Gamma id="Gamma.2" name="distr">
                        <parameter id="RealParameter.5" spec="parameter.RealParameter" estimate="false" name="alpha">0.05</parameter>
                        <parameter id="RealParameter.6" spec="parameter.RealParameter" estimate="false" name="beta">10.0</parameter>
                    </Gamma>
                </prior>
                <prior id="RateCGPrior.s:{tag}" name="distribution" x="@rateCG.s:{tag}">
                    <Gamma id="Gamma.3" name="distr">
                        <parameter id="RealParameter.7" spec="parameter.RealParameter" estimate="false" name="alpha">0.05</parameter>
                        <parameter id="RealParameter.8" spec="parameter.RealParameter" estimate="false" name="beta">10.0</parameter>
                    </Gamma>
                </prior>
                <prior id="RateGTPrior.s:{tag}" name="distribution" x="@rateGT.s:{tag}">
                    <Gamma id="Gamma.5" name="distr">
                        <parameter id="RealParameter.11" spec="parameter.RealParameter" estimate="false" name="alpha">0.05</parameter>
                        <parameter id="RealParameter.12" spec="parameter.RealParameter" estimate="false" name="beta">10.0</parameter>
                    </Gamma>
                </prior>
                <prior id="rhoPrior.t:{tag}" name="distribution" x="@rho.t:{tag}">
                    <OneOnX id="OneOnX.0" name="distr"/>
                </prior>
                <prior id="popSizePrior.t:{tag}" name="distribution" x="@popSize.t:{tag}">
                    <OneOnX id="OneOnX.4" name="distr"/>
                </prior>
            </distribution>
            <distribution id="likelihood" spec="CompoundDistribution">
                <distribution id="treeLikelihood.{tag}" spec="bacter.model.ACGLikelihood" data="@{tag}" locus="@locus.{tag}" tree="@Tree.t:{tag}">
                    <siteModel id="SiteModel.s:{tag}" spec="SiteModel">
                        <parameter id="mutationRate.s:{tag}" spec="parameter.RealParameter" estimate="false" name="mutationRate">1.0</parameter>
                        <parameter id="gammaShape.s:{tag}" spec="parameter.RealParameter" estimate="false" name="shape">1.0</parameter>
                        <parameter id="proportionInvariant.s:{tag}" spec="parameter.RealParameter" estimate="false" lower="0.0" name="proportionInvariant" upper="1.0">0.0</parameter>
                        <substModel id="gtr.s:{tag}" spec="GTR" rateAC="@rateAC.s:{tag}" rateAG="@rateAG.s:{tag}" rateAT="@rateAT.s:{tag}" rateCG="@rateCG.s:{tag}" rateGT="@rateGT.s:{tag}">
                            <parameter id="rateCT.s:{tag}" spec="parameter.RealParameter" estimate="false" lower="0.0" name="rateCT">1.0</parameter>
                            <frequencies id="estimatedFreqs.s:{tag}" spec="Frequencies" frequencies="@freqParameter.s:{tag}"/>
                        </substModel>
                    </siteModel>
                    <branchRateModel id="StrictClock.c:{tag}" spec="beast.base.evolution.branchratemodel.StrictClockModel">
                        <parameter id="clockRate.c:{tag}" spec="parameter.RealParameter" estimate="false" name="clock.rate">1.0</parameter>
                    </branchRateModel>
                </distribution>
            </distribution>
        </distribution>
        <operator id="addRemove_popModelConstant.t:{tag}" spec="bacter.operators.AddRemoveConversion" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" weight="10.0"/>
        <operator id="addRemoveDetour_popModelConstant.t:{tag}" spec="bacter.operators.AddRemoveDetour" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" weight="10.0"/>
        <operator id="CFUniform_popModelConstant.t:{tag}" spec="bacter.operators.CFUniform" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="CFWB_popModelConstant.t:{tag}" spec="bacter.operators.CFWilsonBalding" acg="@Tree.t:{tag}" alpha="0.1" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="CFCS_popModelConstant.t:{tag}" spec="bacter.operators.CFConversionSwap" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="CFSTS_popModelConstant.t:{tag}" spec="bacter.operators.CFSubtreeSlide" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="CFSTXnarrow_popModelConstant.t:{tag}" spec="bacter.operators.CFSubtreeExchange" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="CFSTXwide_popModelConstant.t:{tag}" spec="bacter.operators.CFSubtreeExchange" acg="@Tree.t:{tag}" delta="@delta.t:{tag}" isNarrow="false" populationModel="@popModelConstant.t:{tag}" rho="@rho.t:{tag}" weight="10.0"/>
        <operator id="ACGscale.t:{tag}" spec="bacter.operators.ACGScaler" acg="@Tree.t:{tag}" scaleFactor="0.8" weight="1.0"/>
        <operator id="addRemoveRedundant.t:{tag}" spec="bacter.operators.AddRemoveRedundantConversion" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CRswap.t:{tag}" spec="bacter.operators.ConvertedRegionSwap" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CRshift.t:{tag}" spec="bacter.operators.ConvertedRegionShift" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CRBshift.t:{tag}" spec="bacter.operators.ConvertedRegionBoundaryShift" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="mergeSplit.t:{tag}" spec="bacter.operators.MergeSplitConversion" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CEhop.t:{tag}" spec="bacter.operators.ConvertedEdgeHop" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CEflip.t:{tag}" spec="bacter.operators.ConvertedEdgeFlip" acg="@Tree.t:{tag}" weight="1.0"/>
        <operator id="CEslide.t:{tag}" spec="bacter.operators.ConvertedEdgeSlide" acg="@Tree.t:{tag}" apertureSize="0.05" weight="1.0"/>
        <operator id="RateACScaler.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="rateAC.s:{tag}"/>
            <operator id="AVMNOperator.{tag}" spec="kernel.AdaptableVarianceMultivariateNormalOperator" allowNonsense="true" beta="0.05" burnin="400" initial="800" weight="0.1">
                <transformations id="AVMNSumTransform.{tag}" spec="operator.kernel.Transform$LogConstrainedSumTransform">
                    <f idref="freqParameter.s:{tag}"/>
                </transformations>
                <transformations id="AVMNLogTransform.{tag}" spec="operator.kernel.Transform$LogTransform">
                    <f idref="rateAC.s:{tag}"/>
                    <f idref="rateAG.s:{tag}"/>
                    <f idref="rateAT.s:{tag}"/>
                    <f idref="rateCG.s:{tag}"/>
                    <f idref="rateGT.s:{tag}"/>
                </transformations>
                <transformations id="AVMNNoTransform.{tag}" spec="operator.kernel.Transform$NoTransform"/>
            </operator>
            <operator id="RateACScalerX.s:{tag}" spec="kernel.BactrianScaleOperator" parameter="@rateAC.s:{tag}" scaleFactor="0.5" upper="10.0" weight="0.1"/>
        </operator>
        <operator id="RateAGScaler.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="rateAG.s:{tag}"/>
            <operator idref="AVMNOperator.{tag}"/>
            <operator id="RateAGScalerX.s:{tag}" spec="kernel.BactrianScaleOperator" parameter="@rateAG.s:{tag}" scaleFactor="0.5" upper="10.0" weight="0.1"/>
        </operator>
        <operator id="RateATScaler.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="rateAT.s:{tag}"/>
            <operator idref="AVMNOperator.{tag}"/>
            <operator id="RateATScalerX.s:{tag}" spec="kernel.BactrianScaleOperator" parameter="@rateAT.s:{tag}" scaleFactor="0.5" upper="10.0" weight="0.1"/>
        </operator>
        <operator id="RateCGScaler.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="rateCG.s:{tag}"/>
            <operator idref="AVMNOperator.{tag}"/>
            <operator id="RateCGScalerX.s:{tag}" spec="kernel.BactrianScaleOperator" parameter="@rateCG.s:{tag}" scaleFactor="0.5" upper="10.0" weight="0.1"/>
        </operator>
        <operator id="RateGTScaler.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="rateGT.s:{tag}"/>
            <operator idref="AVMNOperator.{tag}"/>
            <operator id="RateGTScalerX.s:{tag}" spec="kernel.BactrianScaleOperator" parameter="@rateGT.s:{tag}" scaleFactor="0.5" upper="10.0" weight="0.1"/>
        </operator>
        <operator id="FrequenciesExchanger.s:{tag}" spec="AdaptableOperatorSampler" weight="0.05">
            <parameter idref="freqParameter.s:{tag}"/>
            <operator idref="AVMNOperator.{tag}"/>
            <operator id="FrequenciesExchangerX.s:{tag}" spec="operator.kernel.BactrianDeltaExchangeOperator" delta="0.01" weight="0.1">
                <parameter idref="freqParameter.s:{tag}"/>
            </operator>
        </operator>
        <operator id="popSizeScaler.t:{tag}" spec="ScaleOperator" parameter="@popSize.t:{tag}" scaleFactor="0.8" weight="1.0"/>
        <operator id="rhoScaler.t:{tag}" spec="ScaleOperator" parameter="@rho.t:{tag}" scaleFactor="0.8" weight="1.0"/>
        <operator id="ACGupDown3.t:{tag}" spec="bacter.operators.ACGScaler" acg="@Tree.t:{tag}" scaleFactor="0.8" weight="1.0">
            <parameterInverse idref="rho.t:{tag}"/>
        </operator>
        <operator id="deltaScaler.t:{tag}" spec="ScaleOperator" parameter="@delta.t:{tag}" scaleFactor="0.8" weight="1.0"/>
        <logger id="tracelog" spec="Logger" fileName="bacter.log" logEvery="1000" sanitiseHeaders="true">
            <log idref="posterior"/>
            <log idref="likelihood"/>
            <log idref="prior"/>
            <log id="ACGStatsLog.t:{tag}" spec="bacter.util.ConversionGraphStatsLogger" acg="@Tree.t:{tag}"/>
            <log idref="rateAC.s:{tag}"/>
            <log idref="rateAG.s:{tag}"/>
            <log idref="rateAT.s:{tag}"/>
            <log idref="rateCG.s:{tag}"/>
            <log idref="rateGT.s:{tag}"/>
            <log idref="freqParameter.s:{tag}"/>
            <log idref="popSize.t:{tag}"/>
            <log idref="rho.t:{tag}"/>
            <log idref="delta.t:{tag}"/>
        </logger>
        <logger id="screenlog" spec="Logger" fileName="bacter" logEvery="1000">
            <log id="ESS.0" spec="util.ESS" arg="@posterior"/>
            <log id="ESS.1" spec="util.ESS" arg="@rho.t:{tag}"/>
            <log id="ESS.2" spec="util.ESS" arg="@delta.t:{tag}"/>
        </logger>
        <logger id="ACGlogger.t:{tag}" spec="Logger" fileName="bacter.trees" logEvery="10000" mode="tree">
            <log idref="Tree.t:{tag}"/>
        </logger>
    </run>

</beast>"""

    # ------------------------------------------------------------------
    # 3. Guardar el resultado
    # ------------------------------------------------------------------
    xml_out.write_text(template, encoding="utf-8")

def read_beast_log(path):
    with open(path) as f:
        lines = [l.rstrip() for l in f if not l.startswith("#")]
    if not lines:
        sys.exit(f"ERROR: {path} empty")
    header = lines[0].split()
    data = [l.split() for l in lines[1:] if l.strip()]
    df = _pd.DataFrame(data, columns=header).apply(_pd.to_numeric, errors="ignore")
    return df

def autocorr(x):
    x = _np.asarray(x)
    x = x - x.mean()
    n = len(x)
    corr = _np.correlate(x, x, mode="full")[n-1:] / (x.var() * n)
    return corr

def effective_sample_size(x):
    acf = autocorr(x)
    neg = _np.where(acf[1:] < 0)[0]
    M = neg[0] + 1 if len(neg)>0 else len(acf)
    tau = 1.0 + 2.0 * acf[1:M].sum()
    return len(x) / tau

def compute_ess(df, burnin, prefixes=("rho", "delta")):
    N = len(df)
    start = int(N * burnin) if 0<burnin<1 else int(burnin)
    sub = df.iloc[start:].reset_index(drop=True)
    out = {}
    for p in prefixes:
        cols = [c for c in sub.columns if c.startswith(p)]
        out[p] = None if not cols else effective_sample_size(sub[cols[0]].values)
    return out

# ------------------------------------------------------------------
# The new paths are resolved DYNAMICALLY inside
#   results/<tag>/santa/
# We initialize to None to avoid accidental access outside
# the experiment directory.
# ------------------------------------------------------------------
# ------------------------------------------------------------------
#  FILTERING sequence_events_map  ➜   reusable function
# ------------------------------------------------------------------

def filter_sequence_event_map(infile: Path, outfile: Path) -> None:
    """
    Creates *outfile* from *infile* keeping:
        • all R:id@ tokens
        • P:id@… tokens only if that id appears as R: in
          any other sequence **and** does NOT appear as R: in the line.
    """
    lines = [ln.rstrip("\n") for ln in infile.read_text().splitlines()]
    if not lines:
        raise ValueError(f"{infile} is empty")

    header, *seq_lines = lines
    token_re = re.compile(r"[RP]:(\d+)@")

    # --- ids that are R: at some point ---
    R_ids: set[str] = set()
    for ln in seq_lines:
        rhs = ln.split("*", 1)[1]
        for tok in rhs.strip("[]").split(" ; "):
            if tok.startswith("R:"):
                m = token_re.match(tok)
                if m:
                    R_ids.add(m.group(1))

    # --- filter each line ---
    out: list[str] = [header]
    for ln in seq_lines:
        seq_id, raw = ln.split("*", 1)
        tokens = [t.strip() for t in raw.strip("[]").split(" ; ") if t.strip()]
        keep, local_R = [], set()
        for tok in tokens:
            if tok.startswith("R:"):
                local_R.add(token_re.match(tok).group(1))
                keep.append(tok)
        for tok in tokens:
            if tok.startswith("P:"):
                ev_id = token_re.match(tok).group(1)
                if ev_id in R_ids and ev_id not in local_R:
                    keep.append(tok)
        out.append(f"{seq_id}*[{' ; '.join(keep)}]")

    outfile.write_text("\n".join(out) + "\n")
    print(green(f"✓ Written leaked map → {outfile}"))

# These files now live inside results/<tag>/santa.
# They are assigned at runtime inside `parse_santasim`.
REC_EVENTS_FILE : Path | None = None
SEQ_EVENTS_FILE : Path | None = None
Interval = Tuple[int, int]            # (start, end) inclusive
TAG_RX   = re.compile(r"R:(\d+)")
ROOT_ID  = -1                          # ancestral genome label

# ---------------- utilidades de rangos --------------------------------------
def parse(txt: str) -> List[Interval]:
    return [] if not txt else [tuple(map(int, r.split("-"))) for r in txt.split(",")]

def merge(ints: List[Interval]) -> List[Interval]:
    if not ints: return []
    ints.sort()
    out = [list(ints[0])]
    for s, e in ints[1:]:
        ls, le = out[-1]
        if s <= le + 1:
            out[-1][1] = max(le, e)
        else:
            out.append([s, e])
    return [tuple(x) for x in out]

def intersect(a: List[Interval], b: List[Interval]) -> List[Interval]:
    return merge([(max(s1, s2), min(e1, e2))
                  for s1, e1 in a for s2, e2 in b if max(s1, s2) <= min(e1, e2)])

def unite(a: List[Interval], b: List[Interval]) -> List[Interval]:
    return merge(a + b)

def to_str(ints: List[Interval]) -> str:
    return ",".join(f"{s}-{e}" for s, e in ints)
# ──────────── BACK-COMPAT HELPERS ────────────
# The original compute_detectability block was still using
# the old aliases `_merge` and `_to_str`. We create simple
# wrappers that delegate to the new versions.
def _merge(ints: List[Interval]) -> List[Interval]:
    """Backward-compatible alias that forwards to ``merge``."""
    return merge(ints)

def _to_str(ints: List[Interval]) -> str:
    """Backward-compatible alias that forwards to ``to_str``."""
    return to_str(ints)
# ─────────────────────────────────────────────

# ---------------- loading rec_events.txt -----------------------------------
def load_rec(path: Path) -> Dict[int, dict]:
    rec = {}
    with path.open() as f:
        _ = f.readline()
        for ln in f:
            ev, gen, p1f, p1t, p2f, p2t = (ln.rstrip("\n").split("*") + ["", "", ""])[:6]
            ev = int(ev)
            rec[ev] = {
                "Gen": int(gen) if gen else None,
                "P1_frags": parse(p1f),
                "P1_tags": p1t or None,
                "P2_frags": parse(p2f),
                "P2_tags": p2t or None,
            }
    return rec


# -------------- loading sequence_events_map_filtered.txt -------------------
def load_seq(path: Path) -> Dict[int, dict]:
    seq = {}
    with path.open() as f:
        _ = f.readline()
        for ln in f:
            ln = ln.strip()
            if not ln: continue
            sid, evs = ln.split("*", 1)
            sid = int(sid)
            toks = [t.strip() for t in evs.strip("[]").split(";") if t.strip()]
            R, P = [], {}
            for t in toks:
                if t.startswith("R:"):
                    R.append(int(t[2:].split("@", 1)[0]))
                elif t.startswith("P:"):
                    ev, g, rg = t[2:].rstrip("]").split("@", 2)
                    P[int(ev)] = {"Gen": int(g), "Region": rg}
            seq[sid] = {"R": R, "P": P}
    return seq


# ------------- immediate parent map by side --------------------------------
def parent_map(R: List[int], rec: Dict[int, dict]) -> Dict[int, Dict[str, int]]:
    idx = {ev: i for i, ev in enumerate(R)}
    parent = {ev: {"P1": None, "P2": None} for ev in R}

    print(f"[parent_map] === sequence with events {R}")

    for child in R[1:]:          # the first one never has an ancestor
        for side in ("P1", "P2"):
            tags = rec[child][f"{side}_tags"]
            if not tags:
                continue
            cands = []
            for tag in tags.split(";"):
                m = TAG_RX.search(tag)
                if m:
                    anc = int(m.group(1))
                    if anc in idx and idx[anc] < idx[child]:
                        cands.append(anc)
            if cands:
                # the closest ancestor (highest index < child)
                parent[child][side] = max(cands, key=lambda e: idx[e])

    print(f"[parent_map] → result: {parent}\n")

    return parent


# ------------- detectability propagation (respects nesting) ----------
def detectability(R: List[int], rec: Dict[int, dict]) -> Dict[int, Dict[str, List[Interval]]]:
    # ------------------------------------------------------------------
    #   1) Normalize the keys of the «rec» dictionary
    #      (some parsing paths leave them as str)
    # ------------------------------------------------------------------
    if rec and isinstance(next(iter(rec.keys())), str):
        rec = {int(k): v for k, v in rec.items()}

    det = {ev: {"P1": [], "P2": []} for ev in R}

    # ────────────────────────────────────────────────────────────
    #  TRIVIAL CASE
    #  • There is only ONE event in the sequence  (len(R) == 1)
    #  • The event is NOT annotated as parent of any other
    #    (P1_tags/P2_tags empty ➜ parent_map would return None)
    #
    #  In that scenario there are no "child" events covering fragments,
    #  so EVERYTHING inherited from each parental remains visible.
    #  We return the original ranges as-is and finish.
    # ────────────────────────────────────────────────────────────
    if len(R) == 1:
        ev = R[0]
        det[ev]['P1'] = rec.get(ev, {}).get('P1_frags', [])
        det[ev]['P2'] = rec.get(ev, {}).get('P2_frags', [])
        return det

    if not R:                      # sequence without recombinations
        return det

    parent = parent_map(R, rec)

    last = R[-1]


    det[last]["P1"], det[last]["P2"] = rec[last]["P1_frags"], rec[last]["P2_frags"]
    print(f"[detectability] ▼ start for sequence {R}")
    print(f"  final event {last}: P1={det[last]['P1']}  P2={det[last]['P2']}")



    for child in reversed(R):
        for side in ("P1", "P2"):
            rngs = det[child][side]
            if not rngs:
                continue
            anc = parent[child][side]
            if anc is None:
                continue

            # the inherited piece is distributed by intersection with ancestor frags
            before = (det[anc]["P1"][:], det[anc]["P2"][:])
            det[anc]["P1"] = unite(det[anc]["P1"],
                                   intersect(rngs, rec[anc]["P1_frags"]))
            det[anc]["P2"] = unite(det[anc]["P2"],
                                   intersect(rngs, rec[anc]["P2_frags"]))
            print(f"    propagate {child}->{anc} ({side})  "
                  f"before={before}  now={(det[anc]['P1'], det[anc]['P2'])}")


    return det

def compute_detectability(rec: dict[int, dict],          # dict with P1_frags / P2_frags
                          seq_event_map: dict[str, str]) -> dict[int, dict[str,str]]:

    """
    Returns {ev_id: {'P1': 'a-b,c-d'  or  'ND',
                     'P2': 'a-b,c-d'  or  'ND'}}
    """
    # 1) re-order events by sequence exactly like
    #    detect_visibles_final.py
    seq_data = {}                     # <- condensed copy of SEQ_EVENTS_FILE
    for sid, tags in seq_event_map.items():
        R = [int(ev) for ev in re.findall(r"R:(\d+)", tags)]
        seq_data[int(sid)] = {"R": sorted(R), "P": {}}   # 'P' not used here

    # 2) we execute the detectability algorithm per sequence
    # ➊ initialize with "raw" fragments; if the propagation
    #   algorithm doesn't overwrite them → they remain as-is (non-nested case)
    # ------------------------------------------------------------------
    # Now we store **all** versions that are observed.
    # Each side (P1/P2) is a *set* of strings:
    #        {'0-990', 'ND', '0-990,1500-2000', …}
    # This way it's enough to do «add()» to accumulate values
    # without exact duplicates.
    # ------------------------------------------------------------------
    det_by_ev = {
        ev: {
            "P1": { _to_str(rec[ev]["P1_frags"]) or "ND" },
            "P2": { _to_str(rec[ev]["P2_frags"]) or "ND" },
        }
        for ev in rec
    }
    for sid, seq in seq_data.items():
        R = seq["R"]
        if not R:
            continue
        det = detectability(R, rec)   # ← use your rec dict
        print(f"[compute_detectability] detectability() result for sec {sid}:\n{det}")
        pprint.pprint(det)

        for ev in R:
            # ➋ we merge what's already visible with what's propagated; if there's no propagation
            #   (event without children) the result remains the original
            # We add the version observed in this sequence
            det_by_ev[ev]["P1"].add(
                _to_str(_merge(det[ev]["P1"])) or "ND"
            )
            det_by_ev[ev]["P2"].add(
                _to_str(_merge(det[ev]["P2"])) or "ND"
            )

    #print("[compute_detectability] --- global accumulation after processing sequence ---")
    #pprint.pprint(det_by_ev)

    # 3) final format 'a-b,c-d' / 'ND'
    out = {}
    # ------------------------------------------------------------------
    # We convert the *sets* into strings joined by "/".
    # Ex.:  {'0-990', 'ND'}  →  '0-990/ND'
    # They are sorted so results are reproducible.
    # ------------------------------------------------------------------
    for ev, d in det_by_ev.items():
        out[ev] = {
            "P1": "/".join(sorted(d["P1"])),
            "P2": "/".join(sorted(d["P2"])),
        }

    #print("[compute_detectability] === final dict to be returned ===")
    #pprint.pprint(out);  print()

    return out

def parse_santasim(events_file: Path,
                   map_file: Path,
                   alignment_file: Path):

    # All paths relative to the experiment's 'santa' folder
    santa_dir = map_file.parent

    # Update global variables for utilities that need them
    global REC_EVENTS_FILE, SEQ_EVENTS_FILE
    REC_EVENTS_FILE = santa_dir / "rec_events.txt"
    SEQ_EVENTS_FILE = santa_dir / "sequence_events_map_filtered.txt"

    # ------------------------------------------------------------------
    # 1.  Generate **once** the filtered map for this replicate
    # ------------------------------------------------------------------
    if not SEQ_EVENTS_FILE.exists():
        filter_sequence_event_map(map_file, SEQ_EVENTS_FILE)

    santa_events = {}

    # ------------------------------------------------------------------
    #  NEW:  read rec_events.txt to get fragments/tags
    # ------------------------------------------------------------------
    rec_events_path = santa_dir / "rec_events.txt"
    if not rec_events_path.exists():
        sys.exit(f"ERROR: {rec_events_path} not found. "
                 "Did you run Santa-Sim with --outputRecoEvents?")

    raw_rec = load_rec(rec_events_path)         # ← returns genuine lists

    # 1)  Old-compat format («Par1_fragments», …) ——>
    all_rec_events = {
        ev: {
            "Par1_fragments": to_str(d["P1_frags"]),
            "Par1_tags"    : d["P1_tags"] or "",
            "Par2_fragments": to_str(d["P2_frags"]),
            "Par2_tags"    : d["P2_tags"] or "",
        }
        for ev, d in raw_rec.items()
    }

    # 2)  Dictionary ready for detectability() ——>
    rec_compat = {
        ev: {
            "P1_frags": d["P1_frags"],
            "P1_tags" : d["P1_tags"] or None,
            "P2_frags": d["P2_frags"],
            "P2_tags" : d["P2_tags"] or None,
        }
        for ev, d in raw_rec.items()
    }

    if not rec_compat:
        raise RuntimeError("rec_compat ended up empty: "
                           "check that rec_events.txt contains data")
    map_dict = {}
    parents_by_event = {}

    with events_file.open() as ef, map_file.open() as mf, alignment_file.open() as af:

        next(ef)
        for line in ef:                      # 2) Iterate over remaining lines
            parts = line.strip().split('*')  # 3) Split by *

            if len(parts) < 5:               # Safety: malformed line → ignore it
                continue

            # We always use numeric event IDs to avoid
            # mismatches between str ↔ int in the rest of the code
            name           = int(parts[0].strip())   # ← event id
            generation      = parts[2]       # Third field (index 2) → "Generation"
            anc_rec_seq     = parts[3]       # Fourth field → "anc_rec_seq"

            # 4) Process fifth field "[XXX,YYY]"
            raw_list        = parts[4].strip('[]')          # remove brackets
            anc_rep_par1, anc_rep_par2 = [x.strip() for x in raw_list.split(',')]
            #print(anc_rep_par1)
            #print(f"Ancestral sequence lengths: {len(anc_rep_par1)},{len(anc_rep_par2)},{len(anc_rec_seq)}")

            # 5) Save everything in the dictionary
            santa_events[name] = {
                "Generation":    generation,
                "anc_rec_seq":   anc_rec_seq,
                "anc_rep_par1":  anc_rep_par1,
                "anc_rep_par2":  anc_rep_par2,
            }

# ── build the dictionary needed by detectability() ────────────────
        next(mf)  # Skip header


        for line in mf:
            line = line.strip()
            if not line:
                continue  # Skip empty lines

            try:
                seq, tag_seq = line.split('*')
            except ValueError:
                continue  # Malformed line

            seq = seq.strip()
            tag_seq = tag_seq.strip().strip('[]')  # remove outer brackets
            tags = [tag.strip() for tag in tag_seq.split(';') if tag.strip()]

            for tag in tags:
                try:
                    tipo, resto = tag.split(':', 1)  # tipo = "P" or "R"
                    parts = resto.split('@')
                    if len(parts) < 2:
                        continue  # incorrect format

                    event_id = int(parts[0].strip())  # the event number

                    if tipo == "P":
                        region = parts[2] if len(parts) > 2 else None
                        if not region:
                            continue

                        if event_id not in parents_by_event:
                            parents_by_event[event_id] = {}

                        if seq not in parents_by_event[event_id]:
                            parents_by_event[event_id][seq] = []

                        parents_by_event[event_id][seq].append(region)

                    elif tipo == "R":
                        if event_id not in parents_by_event:
                            parents_by_event[event_id] = {}

                        if "recombinants" not in parents_by_event[event_id]:
                            parents_by_event[event_id]["recombinants"] = set()

                        parents_by_event[event_id]["recombinants"].add(seq)

                except Exception as e:
                    print(f"Error processing tag '{tag}' in line '{line}': {e}")

        for seq, fragments in map_dict.items():
            for p_key, p_value in fragments.items():
                try:
                    p_value = p_value.strip()
                    _, fragment_info = p_value.split("P:")  # Quita el prefijo 'P:'
                    parts = fragment_info.split("@")

                    donor_id = parts[0]        # por ejemplo '3530'
                    region = parts[2] if len(parts) > 2 else None  # por ejemplo '9729-13401'

                    if region:
                        if donor_id not in parents_by_event:
                            parents_by_event[donor_id] = {}

                        if seq not in parents_by_event[donor_id]:
                            parents_by_event[donor_id][seq] = []

                        parents_by_event[donor_id][seq].append(region)

                except Exception as e:
                    print(f"Error processing {p_value}: {e}")


        for name, evt in santa_events.items():
            # Add info from all_rec_events
            extra = all_rec_events.get(name, {})
            par1_fragments = extra.get("Par1_fragments", "")
            par1_tags = extra.get("Par1_tags", "")
            par2_fragments = extra.get("Par2_fragments", "")
            par2_tags = extra.get("Par2_tags", "")

            santa_events[name].update({
                "Par1_fragments": par1_fragments,
                "Par1_tags": par1_tags,
                "Par2_fragments": par2_fragments,
                "Par2_tags": par2_tags,
                # ─ breakpoints (heuristic: 1st and last frag. of Par2;
                #   if Par2 is empty, we use Par1) ─
                **({
                    "break_ini": parse(par2_fragments)[0][0],
                    "break_end": parse(par2_fragments)[-1][1]
                } if par2_fragments else {
                    "break_ini": parse(par1_fragments)[0][0] if par1_fragments else None,
                    "break_end": parse(par1_fragments)[-1][1] if par1_fragments else None
                })
            })

            par1_seqs = set()
            par2_seqs = set()

            for event_id in santa_events:
                if event_id in parents_by_event:
                    par1_seqs = set()
                    par2_seqs = set()
                    recombinants = sorted(list(parents_by_event[event_id].get("recombinants", [])))

                    for seq_id, regions in parents_by_event[event_id].items():
                        if seq_id == "recombinants":
                            continue
                        for region in regions:
                            try:
                                start = int(region.split('-')[0])
                                if start == 0:
                                    par1_seqs.add(seq_id)
                                else:
                                    par2_seqs.add(seq_id)
                            except Exception as e:
                                print(f"Error processing region '{region}' for event '{event_id}' in sequence '{seq_id}': {e}")

                    santa_events[event_id]["par1_seqs"] = sorted(par1_seqs)
                    santa_events[event_id]["par2_seqs"] = sorted(par2_seqs)
                    santa_events[event_id]["recombinants"] = recombinants
                    santa_events[event_id]["recombinants_num"]  = [int(x) for x in recombinants]


        alignment = list(SeqIO.parse(af, "fasta"))

        #print(f"ALINEAMIENTO: {alignment}")

        # Cargar el alineamiento como diccionario id → secuencia
        seq_dict = {record.id: str(record.seq) for record in alignment}

        def calcular_similitud(seq1, seq2):
            # Las secuencias deben tener igual longitud
            if len(seq1) != len(seq2):
                raise ValueError(f"Longitudes distintas – {len(seq1)} != {len(seq2)}")
                return None
            diferencias = sum(1 for a, b in zip(seq1, seq2) if a != b)
            return (len(seq1) - diferencias) / len(seq1)

        for event_id in santa_events:
            anc_seq_1 = santa_events[event_id]["anc_rep_par1"]
            anc_seq_2 = santa_events[event_id]["anc_rep_par2"]

            #print(f"EVENTO {event_id}:")

            for label, anc_seq in [("PAR1", anc_seq_1), ("PAR2", anc_seq_2)]:
                if not anc_seq:
                    print(f"  {label}: [no ancestral sequence]")
                    santa_events[event_id][f"Sim__seq_{label}"] = []
                    continue

                resultados = []
                for target_id, target_seq in seq_dict.items():
                    try:
                        sim = calcular_similitud(anc_seq, target_seq)
                        resultados.append((target_id, sim))
                    except Exception as e:
                        print(f"Error comparing event {event_id}, {label} with {target_id}: {e}")


                # WARNING: WHEN THE ANCESTRAL SEQUENCE PRINTING IS FIXED IN SANTASIM, CHECK IF IT IS NEEDED TO DELETE THE FOLLOWING LINE
                resultados = [(sid, s) for sid, s in resultados if s is not None]

                resultados.sort(key=lambda x: x[1], reverse=True)
                formateado = [f"{seq_id}:{sim*100:.3f}%" for seq_id, sim in resultados]

                #print(f"  {label}: {'  '.join(formateado)}")

                # Add to original dictionary
                santa_events[event_id][f"Sim__seq_{label}"] = formateado
                #print(f"DEBUG: Added Sim__seq_{label} to event {event_id}: {formateado[:3]}...")

        # ------------------------------------------------------------
        # Detectability (which pieces remain visible in the sample)
        # ------------------------------------------------------------
        seq_event_map = load_seq(SEQ_EVENTS_FILE)          # ← REAL map
        
        # ---------------  compatibility patch -----------------
        # 1)  seq_event_map = 'new' output from load_seq()
        # 2)  old_seq_event_map = "plain text" version expected by
        #     compute_detectability()  (uses regexp over strings).
        old_seq_event_map = {}
        for sid, d in seq_event_map.items():
            toks = []
            for ev in d.get("R", []):
                toks.append(f"R:{ev}@")    # the '@' was used by the old format
            for ev, pdata in d.get("P", {}).items():
                toks.append(f"P:{ev}@{pdata['Gen']}@{pdata['Region']}")
            old_seq_event_map[str(sid)] = "; ".join(toks)

        # 3)  DO NOT overwrite rec_compat again!
        #     We already created it above from
        #     parental_tags_by_event.txt maintaining the correct
        #     keys ('P1_frags', 'P2_frags', …).  If we
        #     re-generate it here all information is lost and
        #     everything ends up as 'ND'.

        # 4)  We calculate ONCE with the good dictionary!
        det = compute_detectability(rec_compat, old_seq_event_map)
        for ev_id, d in det.items():
            se = santa_events.get(ev_id)
            if se is None:                 # "ghost" id (e.g. 0)
                #print(f"[INFO] detectability() returned id {ev_id}, "
                #      "which is not in santa_events – skipping.")
                continue

            se['det_par1'] = d['P1']       # '0-7355,12000-15000'  or  'ND'
            se['det_par2'] = d['P2']

        #print(santa_events,"\n\n",all_rec_events,"\n\n",map_dict,"\n\n",parents_by_event,"\n\n")
        #print(santa_events)
        for name, data in santa_events.items():
            
            # Compress specific fields
            anc_rec_seq = data.get("anc_rec_seq", "")
            anc_rep_par1 = data.get("anc_rep_par1", "")
            anc_rep_par2 = data.get("anc_rep_par2", "")

            def short(s):
                return s[:5] + "..." if len(s) > 5 else s

            def short_list(lst, max_items=3):
                if not lst:
                    return "[]"
                if len(lst) <= max_items:
                    return " ".join(lst)
                return " ".join(lst[:max_items]) + " ..."

            print(f"Event: {name}")
            print(f"  Generation     : {data.get('Generation')}")
            print(f"  anc_rec_seq    : {short(anc_rec_seq)}")
            print(f"  anc_rep_par1   : {short(anc_rep_par1)}")
            print(f"  Par1_fragments : {data.get('Par1_fragments')}")
            print(f"  Par1_tags      : {data.get('Par1_tags')}")
            print(f"  par1_seqs      : {data.get('par1_seqs')}")
            print(f"  det_P1        : {data.get('det_par1', 'ND')}")
            print(f"  Sim__seq_Par1  : {short_list(data.get('Sim__seq_PAR1'))}")
            print("– " * 20)
            print(f"  anc_rep_par2   : {short(anc_rep_par2)}")
            print(f"  Par2_fragments : {data.get('Par2_fragments')}")
            print(f"  Par2_tags      : {data.get('Par2_tags')}")
            print(f"  par2_seqs      : {data.get('par2_seqs')}")
            print(f"  det_P2        : {data.get('det_par2', 'ND')}")
            print(f"  recombinants   : {data.get('recombinants')}")
            print(f"  Sim__seq_Par2  : {short_list(data.get('Sim__seq_PAR2'))}")
            print("-" * 40)
        return santa_events

        # From here on I'm checking repeated recombinants to see which events are detectable in the end.

        recomb_to_events = {}

        for event_id, data in santa_events.items():
            recombinants = data.get("recombinants", [])
            for rec in recombinants:
                if rec not in recomb_to_events:
                    recomb_to_events[rec] = []
                recomb_to_events[rec].append(event_id)

        print("\nRecombinant repeats and relevant events:")
        for rec, eventos in recomb_to_events.items():
            if len(eventos) > 1:
                # Sort by numeric ID value
                eventos_ordenados = sorted(eventos)            # already ints → natural order
                eventos_menos_ultimo = [str(e)                 # convert to str
                                         for e in eventos_ordenados[:-1]]
                print(f"Sequence {rec} appears in events (except the last one): "
                      f"{', '.join(map(str, eventos_menos_ultimo))}")

        # Step 1: build inverse index recombinant → events
        recomb_to_events = defaultdict(set)

        for event_id, data in santa_events.items():
            for rec in data.get("recombinants", []):
                recomb_to_events[rec].add(event_id)

        # Step 2: analyze each event
        for event_id, data in santa_events.items():
            recombs = data.get("recombinants", [])
            if not recombs:
                continue

            otros_eventos = []

            for rec in recombs:
                eventos_rec = recomb_to_events.get(rec, set()) - {event_id}
                if not eventos_rec:
                    break  # At least one is not in any other event → we discard
                otros_eventos.extend(eventos_rec)

            else:
                # If we didn't break → all recombs appear in other events
                eventos_unicos = sorted(set(otros_eventos), key=int)

                if len(eventos_unicos) == 1:
                    print(f"All recombinant sequences of the event {event_id} "
                          f"also appear at the event {eventos_unicos[0]}")
                else:
                    eventos_unicos_str = [str(e) for e in eventos_unicos]
                    print(f"The recombinant sequences of the event {event_id} "
                          f"are distributed among the events: "
                          f"{', '.join(eventos_unicos_str)}")

# ───────────────────────── ClonalFrameML ──────────────────────────
def parse_clonalframeml_pair(
        imp_path: str | Path,
        tree_path: str | Path,
        fasta_path: str | Path          # ← NEW mandatory parameter
) -> Tuple[Dict[str, Dict[str, object]], int]:

    imp_path = Path(imp_path)
    tree_path = Path(tree_path)
    fasta_path = Path(fasta_path)

    first_rec = next(SeqIO.parse(fasta_path, "fasta"))
    max_span  = len(first_rec.seq)

    # ------------------------------------------------------------------
    # 1) Load annotated tree:  node_label → list of leaves
    # ------------------------------------------------------------------
    tree = Phylo.read(tree_path, "newick")

    leaves_of: Dict[str, List[str]] = {}
    for clade in tree.find_clades():
        label = clade.name
        if label is None or label.startswith("#"):   # (just in case)
            continue
        if not clade.is_terminal():
            leaves_of[label] = sorted(t.name for t in clade.get_terminals())

    # ------------------------------------------------------------------
    # 2) Go through the .importation_status.txt
    # ------------------------------------------------------------------
    events: Dict[str, Dict[str, object]] = {}
    discarded = 0 
    with imp_path.open() as fh:
        next(fh)                         # skip header "Node Beg End"
        for line in fh:
            node, beg_s, end_s = line.split()
            beg, end = int(beg_s), int(end_s)

            # optional filter
            if beg == 1 and end == max_span:
                discarded += 1 
                continue

            # ── NEW: complementary intervals ──────────────────────────
            #
            #  Detected fragment  =  break_1
            #  Rest of genome     =  break_2  (can be 1 or 2 pieces)
            #
            break_1 = f"{beg}-{end}"

            # ───────── calculate "complementary" zone break_2 ─────────
            if beg == 1 and end == max_span:           # covers all → no complement
                break_2 = ""
            elif beg == 1:                              # cut only at 3'
                break_2 = f"{end+1}-{max_span}"
            elif end == max_span:                       # cut only at 5'
                break_2 = f"1-{beg-1}"
            else:                                       # internal cut (two pieces)
                break_2 = f"1-{beg-1},{end+1}-{max_span}"

            events[node] = {
                "break_ini": beg,
                "break_end": end,
                "break_1" : break_1,
                "break_2" : break_2,
                "sequences": leaves_of.get(node, [])
            }

    return events, discarded

# ──────────────────────────── Bacter ──────────────────────────────

def leaves(clade):
    """'Normal' leaves under a clade (ignores #n nodes)."""
    return [t.name for t in clade.get_terminals() if not t.name.startswith("#")]

def siblings(clade, parent_of):
    """Leaves under the immediate siblings of *clade*."""
    par = parent_of.get(clade)
    outs = []
    if par:
        for sib in par.clades:
            if sib is not clade:
                outs.extend(leaves(sib))
    return outs

# ────────────────────────────────────────────────────────────────────
def parse_bacter_nexus(tree_file: Path, max_ascii: int = 120
                               ) -> Dict[str, Dict[str, List[str]]]:
    """
    Devuelve p.ej.:

        { '#0': { 'group1': [...],
                  'group2': [...],
                  'group3': [...] },
          '#1': { ... } }
    """
    txt = tree_file.read_text()

    # 1) get num→taxlabel map + Newick
    translate, raw_newick = {}, None
    in_tr = False
    for ln in txt.splitlines():
        s = ln.strip()
        if s.lower().startswith("translate"):
            in_tr = True;  s = s[len("translate"):].strip()
        if in_tr:
            if s.endswith(";"):
                in_tr = False;  s = s[:-1]
            if s:
                num, name = s.rstrip(",").split(None, 1)
                translate[num] = name
            continue
        if s.lower().startswith("tree"):
            raw_newick = s.split("=", 1)[1].strip().rstrip(";")
            break
    if raw_newick is None:
        sys.exit("Tree not found in NEXUS")

    # 2) substitute numbers → names
    newick_num = re.sub(r'(?<=\(|,)(\d+)(?=[:\[,)\s])',
                        lambda m: translate.get(m.group(1), m.group(1)),
                        raw_newick)

    # 3) rename duplicates #n so Biopython doesn't complain
    counts, dup_map = defaultdict(int), {}
    def dedup(m):
        tag = m.group(0)
        counts[tag] += 1
        if counts[tag] == 1:
            return tag
        new = f"{tag}__dup{counts[tag]-1}"
        dup_map[new] = tag
        return new
    newick = re.sub(r'#\d+', dedup, newick_num)

    # 4) load the tree
    tree = Phylo.read(io.StringIO(newick), "newick")

    # 5) parent_of for fast navigation
    parent_of = {}
    for cl in tree.find_clades(order="level"):
        for ch in cl.clades:
            parent_of[ch] = cl

    # 6) locate occurrences by label
    occ = defaultdict(list)
    for cl in tree.find_clades():
        if cl.name and cl.name.startswith("#"):
            occ[dup_map.get(cl.name, cl.name)].append(cl)

    out = {}
    for tag, pair in occ.items():
        if len(pair) != 2:
            print(f"[INFO] {tag}: is ignored (appears {len(pair)} ≠ 2 times)")
            continue

        a, b = pair
        is_a_anc_b = any(b is d for d in a.find_clades())
        is_b_anc_a = any(a is d for d in b.find_clades())

        # own leaves of each occurrence
        La, Lb = set(leaves(a)), set(leaves(b))
        Sa, Sb = set(siblings(a, parent_of)), set(siblings(b, parent_of))

        # ╭─ CASE 1 ▸ vertical (nested) ────────────────────────────────╮
        if is_a_anc_b ^ is_b_anc_a:       # XOR → one is ancestor of the other
            upper, lower = (a, b) if is_a_anc_b else (b, a)
            g3 = sorted(leaves(lower) or siblings(lower, parent_of))
            out[tag] = {
                "group1": sorted(siblings(lower, parent_of)),
                "group2": sorted(siblings(upper, parent_of)),
                "group3": g3
            }
            continue
        # ╰──────────────────────────────────────────────────────────────╯

        # ╭─ CASE 2 ▸ horizontal (non-nested) ──────────────────────────╮
        #    ‣ decide recombinant clade by heuristic:
        #       · if one has NO own leaves → the OTHER is the donor
        #       · if both have them → choose the smaller exclusive subset
        # ----------------------------------------------------------------
        if La and not Lb:
            rec, don = a, b
        elif Lb and not La:
            rec, don = b, a
        else:
            diff_a, diff_b = La - Lb, Lb - La
            # if leaves overlap, use the smaller difference
            if diff_a and (not diff_b or len(diff_a) <= len(diff_b)):
                rec, don = a, b
            elif diff_b:
                rec, don = b, a
            else:               # identical sets → pick arbitrarily
                rec, don = (a, b) if len(La) <= len(Lb) else (b, a)

        g3 = leaves(rec)
        if not g3:                          # fallback: use exclusive difference
            g3 = sorted((La - Lb) or (Lb - La))

        out[tag] = {
            "group1": sorted(siblings(rec, parent_of)),
            "group2": sorted(siblings(don, parent_of)),
            "group3": sorted(g3)
        }
    return out

# ──────────────────── Extract only Santa results ────────────────────
def extract_santa_only(run_dir: Path):
    """Extracts only Santa events without requiring cfml/bacter"""
    run_dir = Path(run_dir).resolve()
    
    # Only verify that santa exists
    if not (run_dir / "santa").is_dir():
        return {}
    
    # ── Santa‑Sim (event + map) ──
    # parse_santasim expects two files:
    # 1. events_file: recombination_events_final_sample.txt (specific format)
    # 2. Internally also reads rec_events.txt
    santa_ev  = run_dir / "santa" / "recombination_events_final_sample.txt"
    santa_map = run_dir / "santa" / "sequence_events_map_final_sample.txt"
    santa_aln  = run_dir / "santa" / "final_sample.fasta"
    
    santa_dict = {}
    # Only call parse_santasim if the main events file exists
    # (parse_santasim requires both files: events_file and rec_events.txt)
    if santa_ev.exists() and santa_map.exists():
        try:
            santa_dict = parse_santasim(santa_ev, santa_map, santa_aln) or {}
        except Exception as e:
            # If there's an error parsing (e.g., incorrect format), return empty
            print(f"[WARN] Could not parse Santa events from {run_dir.name}: {e}")
            santa_dict = {}
    # If the main file doesn't exist but rec_events.txt exists, 
    # parse_santasim won't work with the current format
    
    return santa_dict

# ──────────────────── Extract and display results ────────────────────
def extract_recombination_results(run_dir: Path, par_threshold: float):
    run_dir = Path(run_dir).resolve()

    # ── check that it's a valid folder ──────────────────────────
    required = ["santa", "cfml", "bacter"]
    if not all((run_dir / d).is_dir() for d in required):
        raise FileNotFoundError(
            f"'{run_dir}' doesn't seem to be a valid run folder "
            f"(missing {', '.join(required)})"
        )

    # ── Santa‑Sim (event + map) ──
    # We pass the NON-filtered version;  parse_santasim will create
    # sequence_events_map_final_sample_filtered.txt internally
    santa_ev  = run_dir / "santa" / "recombination_events_final_sample.txt"
    santa_map = run_dir / "santa" / "sequence_events_map_final_sample.txt"
    santa_aln  = run_dir / "santa" / "final_sample.fasta"
    santa_dict = {}
    if santa_ev.exists() and santa_map.exists():
        santa_dict = parse_santasim(santa_ev, santa_map, santa_aln) or {}
    # ── ClonalFrameML ──
    tag    = run_dir.name                    # 100_50_NC_045512.2
    imp    = run_dir / "cfml" / f"{tag}.importation_status.txt"
    tree   = run_dir / "cfml" / f"{tag}.labelled_tree.newick"
    if imp.exists() and tree.exists():
        cfml_dict, discarded = parse_clonalframeml_pair( imp, tree, santa_aln)            # ← pasa el FASTA        
        #print(f"[INFO] CFML discarded events: {discarded}")
    else:
        cfml_dict = {}

    # ── Bacter ──
    tree_file = run_dir / "bacter" / "summary.tree"
    bacter_dict = parse_bacter_nexus(tree_file) if tree_file.exists() else {}

    # ── imprimir resumen ─────────────────────────────────────────────
    import pprint
    pp = pprint.PrettyPrinter(indent=2, compact=True)
    print("\n=========== SANTA‑SIM ===========");  pp.pprint(santa_dict)
    print("\n========= ClonalFrameML =========")
    print(f"[INFO] CFML discarded events: {discarded}")
    pp.pprint(cfml_dict)
    print("\n=========== Bacter =============="); pp.pprint(bacter_dict)
    print("=================================\n")

    return santa_dict, cfml_dict, bacter_dict

# ──────────────────────── Auxiliaries for correct RDP exception parsing ─────────────────────────

# ─────────────────── Function to clean names with the ^ flag for comparisons ──────────────────
# ──────────────────────────────────────────────────────────────
#  Helper: normalise every weird sequence-id coming from RDP
# ──────────────────────────────────────────────────────────────
_SEQ_ID_FLAG_RX = re.compile(r'\[[A-Za-z]+\]$')          # “4[P]”, “6[T]”, …
_UNKNOWN_RX     = re.compile(r'Unknown\s*\(\s*(\d+)\s*\)', re.I)

def _norm_seq_id(token: str) -> str:
    """
    Return a **clean** identifier from the variety of formats seen in RDP
    output, applying these rules in order:

      1.  Drop a leading '^' (RDP marks recombinants with it).
      2.  \"Unknown (45)\" → \"45\"          (keep only the number).
      3.  Remove trailing single-letter flags in square brackets
          such as \"4[P]\", \"6[T]\" … → \"4\", \"6\".
      4.  If the remaining string *starts* with digits, return those
          digits; otherwise return the cleaned string itself.

    The caller decides whether '^'-prefixed tokens are kept or skipped
    completely (see *clean_set*).
    """
    tok = token.strip()

    # leading “^”
    if tok.startswith("^"):
        tok = tok[1:]

    # Unknown (n)
    m = _UNKNOWN_RX.fullmatch(tok)
    if m:
        return m.group(1)

    # trailing “[X]” flag
    tok = _SEQ_ID_FLAG_RX.sub("", tok)

    # leading digits, if any
    m = re.match(r'(\d+)', tok)
    return m.group(1) if m else tok

def split_flag(token: str) -> tuple[int, str]:
    """
    Separates a Begin/End field that can come as '1234', '5678*', '90$'…
    Returns a tuple (int_value, flag_str).  flag_str will be '', '~', '*', or '$'.
    """
    m = re.fullmatch(r'(\d+)([~\*\$]?)', token.strip())
    if not m:
        raise ValueError(f"Unexpected token at Begin/End: {token}")
    return int(m.group(1)), m.group(2)

# ──────────────────────── RDP → dictionary ─────────────────────────
def parse_rdp_csv(csv_path: Path, excluded: set[str]):
    """
    Processes a '*.fasta.csv' file generated by RDP and returns:
        { (Begin, End): {
              'recombinant': [seq1, seq2, …],
              'minor'      : [seqX, …],
              'major'      : [seqY, …]
          }, … }
    Begin and End are returned as strings as they appear in the CSV.
    """
    events = {}
    with csv_path.open(newline='') as fh:
        # ─── "No recombination detected" message? ────────────────────
        first = next((ln.strip() for ln in fh if ln.strip()), '')
        if first.lower().startswith('no recombination detected'):
            return {'NRD': True}        # <- special signal for the caller
        fh.seek(0)                      # go back to beginning for normal parsing
        reader = csv.reader(fh)
        header = None

        # 1) locate header with Begin/End/Event Number
        for row in reader:
            if not any(cell.strip() for cell in row):
                continue
            cols = [c.strip() for c in row]
            if {'Begin', 'End', 'Recombinant Sequence(s)',
                'Minor Parental Sequence(s)',
                'Major Parental Sequence(s)',
                'Recombination Event Number'}.issubset(cols):
                header = cols
                break
        if header is None:
            raise ValueError(f"Valid header not found in {csv_path.name}")

        # column indices
        idx_begin = header.index('Begin')
        idx_end   = header.index('End')
        idx_rec   = header.index('Recombinant Sequence(s)')
        idx_minor = header.index('Minor Parental Sequence(s)')
        idx_major = header.index('Major Parental Sequence(s)')

        current_key = None

        # --- 2) go through data rows ------------------------------------
        for raw in reader:
            if not any(cell.strip() for cell in raw):
                continue
            if len(raw) < len(header):
                raw += [''] * (len(header) - len(raw))
            row = [c.strip() for c in raw]

            # If the row brings new Begin/End -> a new event may start
            if row[idx_begin] and row[idx_end]:
                beg_raw, end_raw = row[idx_begin], row[idx_end]

                # separate number and flag
                beg_val, beg_flag = split_flag(beg_raw)
                end_val, end_flag = split_flag(end_raw)

                b_int , e_int = sorted([int(beg_val), int(end_val)])

                # Any flag in the exclusion list? -> skip the event
                if beg_flag in excluded or end_flag in excluded:
                    current_key = None        # discard this entire block
                    continue

                b, e = b_int, e_int
                current_key = (b, e)
                #  we only fill the lists here;  parental breakpoints
                #  will be completed later, when we know
                #  the real genome length.
                events.setdefault(current_key,
                    {'recombinant': [], 'minor': [], 'major': [],
                     # →  we only save the pair of cuts;
                     #    parental break-points will be calculated
                     #    later, when we know the real length.
                     'break_ini': b,
                     'break_end': e})

            # --- add data to the current event, if not discarded ---
            if current_key:
                if row[idx_rec]:
                    events[current_key]['recombinant'].append(
                        _norm_seq_id(row[idx_rec]))
                if row[idx_minor]:
                    events[current_key]['minor'].append(
                        _norm_seq_id(row[idx_minor]))
                if row[idx_major]:
                    events[current_key]['major'].append(
                        _norm_seq_id(row[idx_major]))

    return events



def clean_set(seq_list, skip_caret):
    """
    Build a **set** from *seq_list* applying *conditional* normalisation:

      • If *skip_caret* **is True**:
          – Tokens starting with '^' are **skipped** altogether.
          – The remaining tokens are fully normalised by `_norm_seq_id`.

      • If *skip_caret* **is False**:
          – Tokens keep their original form, except that a leading '^'
            (if present) is stripped off.
          – No further changes are applied (the extra flags such as
            \"[P]\", \"[T]\", or \"Unknown (n)\" are preserved).
    """
    out: set[str] = set()
    for tok in seq_list:
        # honour both the function argument *and* the global user choice
        if tok.startswith("^"):
            if skip_caret or _SKIP_CARET_IN_COMP:
                # completely drop recombinant-flagged tips
                continue
            # otherwise: keep the id but strip the caret
            tok = tok[1:]


        tok_clean = _norm_seq_id(tok) if skip_caret else tok
        out.add(tok_clean)
    return out

def _clean_rdp_lists(ev_dict: dict[str, list[str]],
                     *, skip_caret: bool = True) -> None:
    """
    Normalise **recombinant / minor / major** lists inside *ev_dict*
    using `clean_set()`.  The lists are updated *in-place*.
    """
    for key in ("recombinant", "minor", "major"):
        if key in ev_dict:
            ev_dict[key] = list(clean_set(ev_dict[key], skip_caret))

def _text(root: ET.Element, tag: str) -> str | None:
    """
    Returns the .text of the **first** element whose tag suffix
    (without namespace) matches *tag*.  If it doesn't exist → None.
    """
    for el in root.iter():
        if el.tag.endswith(tag):
            return el.text
    return None


def _xml_find(root: ET.Element, tag: str):
    """
    Returns the text of the first element that *contains* `tag` in its name,
    ignoring namespaces.  None if it doesn't exist.
    """
    for el in root.iter():
        if el.tag.rsplit('}', 1)[-1] == tag:
            return el.text
    return None

def parse_cfg_params(xml_path: Path) -> dict:
    """
    Reads cfg.xml WITHOUT assuming fixed namespaces.  Always returns **str** –
    numeric conversions are done further down where needed.
    """
    root = ET.parse(xml_path).getroot()
    params = {
        "popSize"        : _xml_find(root, "populationSize")      or "",
        "mutationRate"   : _xml_find(root, "mutationRate")        or "",
        "recombProb"     : _xml_find(root, "recombinationProbability") or "",
        "dualInfectProb" : _xml_find(root, "dualInfectionProbability") or "",
        "genCount"       : _xml_find(root, "generationCount")     or "",
        "sampleSize"     : _xml_find(root, "sampleSize")          or "",
    }
    return params

def _break_bounds(frag_str: str) -> tuple[int | None, int | None]:
    """
    Returns (ini, end) from the string 'a-b,c-d'.  If empty → (None, None)
    """
    ints = parse(frag_str)          # reuse your original parse()
    return (ints[0][0], ints[-1][1]) if ints else (None, None)


def parse_rdp_dir(rdp_dir: Path,
                  excluded: set[str],
                  start_idx: int | None = None,
                  end_idx:   int | None = None,
                  only_rdp: bool = False,
                  write_no_detection: bool = False):

    rdp_dir = Path(rdp_dir).resolve()
    if not rdp_dir.is_dir():
        raise FileNotFoundError(f"{rdp_dir} is not a valid directory")

    out = {}
    for csv_file in sorted(rdp_dir.glob("*Recombination events.csv")):
        # ── filter by --start/--end range ───────────────────────────
        try:
            idx_val = int(csv_file.name.split('_', 1)[0])
        except ValueError:
            idx_val = None
        if (start_idx is not None and end_idx is not None
                and idx_val is not None
                and not (start_idx <= idx_val <= end_idx)):
            continue

        evs = parse_rdp_csv(csv_file, excluded)
        out[csv_file.name] = evs

        # summary print
        print(f"\n→ {csv_file.name}:")
        if evs and 'NRD' in evs:
            print("  No recombination detected by RDP")
        else:
            for (b, e), g in evs.items():
                print(f"  ({b}, {e}):")
                print(f"    recombinant: {g['recombinant']}")
                print(f"    minor      : {g['minor']}")
                print(f"    major      : {g['major']}")

        # tag = everything before ".fasta"
        tag = csv_file.name.split('.fasta', 1)[0]

        # integrate results into both Excels
        integrate_rdp_results(tag, evs, only_rdp=only_rdp, write_no_detection=write_no_detection)

    if not out:
        print(f"(No *events.csv files found in {rdp_dir})")

    return out





# =========  NEW HELPERS  (paste after remaining utilities) =========
_SIM_RX = re.compile(r'^(\d+):([\d\.]+)%$')


def append_df_to_excel(df: pd.DataFrame,
                       path: Path = Path("results/recombination_summary.xlsx"),
                       sheet: str = "events_by_tool") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    # 1. Does the workbook already exist?
    if path.exists():
        # ---------------------------------------------------------
        # THIS IS THE KEY:
        #    - opens in append mode  (mode="a")
        #    - if_sheet_exists="overlay" avoids creating "events_by_tool1"
        # ---------------------------------------------------------
        with pd.ExcelWriter(path,
                            engine="openpyxl",
                            mode="a",
                            if_sheet_exists="overlay") as wr:
            # new sheet or already existing?
            if sheet in wr.book.sheetnames:
                prev = _pd.read_excel(path, sheet_name=sheet)
                df   = _pd.concat([prev, df], ignore_index=True)
                df.drop_duplicates(
                    subset=['tag', 'tool', 'event_id'], keep='first',
                    inplace=True)
            df = df.loc[:, ~df.columns.duplicated()]
            df.to_excel(wr, sheet_name=sheet,
                        index=False, header=True, startrow=0)

    else:
        # new workbook → headers are written
        df.to_excel(path, sheet_name=sheet, index=False, header=True)

# ──────────────────────────────────────────────────────────────
#  AUX :: intersection ↔ differences statistics
#      Returns  match,  miss_santa,  miss_tool,  pct
# ──────────────────────────────────────────────────────────────
def _set_stats(S:set[str], T:set[str]) -> tuple[int,int,int,float]:
    m     = len(S & T)
    ms    = len(S - T)          # in Santa but not in the tool
    mt    = len(T - S)          # in the tool but not in Santa
    tot   = m + ms + mt
    pct   = round(100*m/tot,2) if tot else 0.0
    return m, ms, mt, pct

# ──────────────────────────────────────────────────────────────
#  AUX :: best (Dice) overlap between
#         • list of Santa interval lists  (P1 / P2 alternatives)
#         • tool interval            (t_ini, t_end)
#  – Merges each Santa interval list and calculates:
#       · nt_inter        intersection length
#       · pct_dice        2·|∩| / (|S|+|T|)
#       · pct_dice_norm   pct_dice · min(|S|,|T|)/max(|S|,|T|)
#  – Returns the *best* trio (nt, pct, pct_norm)
# ──────────────────────────────────────────────────────────────
def _best_break_overlap(
    santa_alternatives: list[list[tuple[int, int]]],
    t_ini: int,
    t_end: int,
    genome_len: int,
) -> tuple[int, float, int, int]:

    if t_ini is None or t_end is None:
        return 0, 0.0, 0, 0               # ← 4 valores de retorno

    # -------------------------------------------------
    # 1) Generate the two orientations for the tool
    # -------------------------------------------------
    # «t_ini» and «t_end» can come as int or str
    def _as_int_pair(v):
        """
        Accepts:
            · int → (v, v)
            · "a-b" or "…/a-b"   → (a, b)
            · None / "ND" / ''  → (None, None)
        Always returns (ini, end) or (None,None)
        """
        if v is None or (isinstance(v, str) and v.strip().upper() in {"", "ND"}):
            return None, None
        if isinstance(v, int):
            return v, v
        # take the FIRST range with numbers it finds
        m = re.search(r"(\d+)\s*-\s*(\d+)", v)
        if not m:
            return None, None
        return int(m.group(1)), int(m.group(2))

    t_ini, t_end = _as_int_pair(t_ini)[0], _as_int_pair(t_end)[1]
    print(f"[DEBUG] Tool range normalized: {t_ini}-{t_end} (genome_len={genome_len})")
    if t_ini is None or t_end is None:
        return 0, 0.0, 0, 0

    tool_alts: list[list[tuple[int, int]]] = [[(t_ini, t_end)]]
    comp_tool: list[tuple[int, int]] = []
    if t_ini > 1:
        comp_tool.append((1, t_ini - 1))
    if t_end < genome_len:
        comp_tool.append((t_end + 1, genome_len))
    if comp_tool:
        tool_alts.append(comp_tool)
    print("[DEBUG] Set of alternatives for the tool:")
    pprint.pprint(tool_alts)


    # -------------------------------------------------
    # 2) Add complements to Santa if needed
    # -------------------------------------------------
    def _complement(ints: list[tuple[int, int]]) -> list[tuple[int, int]]:
        if not ints:
            return []
        ints = sorted(ints)
        res, cur = [], 1
        for a, b in ints:
            if a > cur:
                res.append((cur, a - 1))
            cur = b + 1
        if cur <= genome_len:
            res.append((cur, genome_len))
        return res

    santa_alts_full = santa_alternatives[:] + [_complement(x) for x in santa_alternatives]

    print("[DEBUG] Set of alternatives for Santa (incl. accessories):")
    pprint.pprint(santa_alts_full)

    # -------------------------------------------------
    # 3) Check the 4 possible scenarios
    # -------------------------------------------------
    best_nt = best_dice = best_lenS = best_lenT = 0

    for s_ints in santa_alts_full:
        len_S = sum(b - a + 1 for a, b in s_ints)
        print(f"\n[DEBUG] ► Santa ints {s_ints}  | len_S={len_S}")
        for t_ints in tool_alts:
            len_T = sum(b - a + 1 for a, b in t_ints)
            inter = 0
            for a, b in s_ints:
                for c, d in t_ints:
                    inter += max(0, min(b, d) - max(a, c) + 1)
            tool_ints = ','.join(map(str, t_ints))            # "[(1, 100)],[(250, 300)]"
            print(f"        ‣ Tool ints {tool_ints:<30} len_T={len_T:>6}  inter={inter:>6}", end='')
            if inter == 0:
                print("   → without overlap")
                continue

            dice = 200 * inter / (len_S + len_T)
            print(f"   dice={dice:6.2f}")

            if dice > best_dice:
                best_nt   = inter
                best_dice = round(dice, 2)
                best_lenS = len_S
                best_lenT = len_T
                print(f"        *** NEW BETTER ***  inter={inter}  dice={best_dice}  "
                      f"len_S={best_lenS}  len_T={best_lenT}")

    print(f"\n[DEBUG] Final result → inter={best_nt}, dice={best_dice}, "
          f"len_S={best_lenS}, len_T={best_lenT}")

    return best_nt, best_dice, best_lenS, best_lenT


def _parse_sim_list(raw: list[str]) -> dict[str,float]:
    """
    Converts ['23:99.983%', '31:99.983%', …] to {'23':0.99983, '31':0.99983…}
    """
    out = {}
    for tok in raw:
        if ':' in tok and '%' in tok:
            sid, pct = tok.split(':')
            out[sid] = float(pct.rstrip('%'))/100
    return out

def sim_prod(sim_map: dict[str,float], santa_set: set[str], tool_set: set[str]) -> float:
    """
    Product of similarities (0–1) for sequences that
    *are in tool_set but not in santa_set*.
    If there are none, returns 1.0.
    """
    extras = tool_set - santa_set
    if not extras:
        return "ND"
    prod = 1.0
    found_any = False
    for sid in extras:
        if sid in sim_map:
            found_any = True
        prod *= sim_map.get(sid, 1.0)
    if prod == 1.0 and not found_any:      # ← new patch
        return "ND"
    return round(prod, 6)

def sim_dict(ev: dict) -> dict[str, float]:
    """
    Returns {seq_id : sim_decimal} from Sim__seq_PAR1 / PAR2.
    """
    out = {}
    for key in ("Sim__seq_PAR1", "Sim__seq_PAR2"):
        for tok in ev.get(key, []):
            m = _SIM_RX.match(tok)
            if m:
                out[m.group(1)] = float(m.group(2)) / 100.0
    return out
# ---------------------------------------------------------------------------
# ➊  "event × tool" ROWS  (Santa vs CFML/Bacter/RDP)
# ---------------------------------------------------------------------------
def build_event_rows_v3(tag: str,
                        run_params: dict,
                        santa: dict[int, dict],
                        cfml: dict[str, dict],
                        bacter: dict[str, dict],
                        rdp: dict | None = None
                       ) -> list[dict]:
    """
    Returns a list of rows:
        • ONE for each combination (Santa event × tool event)
        • Each tool group is ordered by % overlap
    If a Santa event has no match in the tool → a
    single row is created with 'ND' in all external columns.
    """
    # ------- internal helpers ----------------------------------------
    def interval_overlap(a1,a2,b1,b2):            # nt of overlap
        return max(0, min(a2,b2) - max(a1,b1) + 1)

    def santa_span(ev):
        return (ev['break_end'] - ev['break_ini'] + 1
                if ev.get('break_ini') and ev.get('break_end') else None)
    def overlap_pct_union(a1: int, a2: int, b1: int, b2: int) -> float:
        """
        % overlap based on UNION SET LENGTH.
        Formula:   2·|A∩B| / (|A| + |B|)    →  0-100 %
        """
        if any(x is None for x in (a1, a2, b1, b2)):
            return 0.0
        inter = max(0, min(a2, b2) - max(a1, b1) + 1)
        if not inter:
            return 0.0
        union_len = (a2 - a1 + 1) + (b2 - b1 + 1) - inter
        return 100.0 * (2 * inter) / ( (a2 - a1 + 1) + (b2 - b1 + 1) )

    # -----------------------------------------------------------------
    #  LONGITUD DEL GENOMA  (para el Dice adj. por probabilidad-azar)
    # -----------------------------------------------------------------
    genome_len = max(ev.get('break_end', 0) or 0 for ev in santa.values())

    # ------- plantilla “base” por evento Santa -----------------------
    base_for_event = {}
    for ev_id, ev in santa.items():
        base_for_event[ev_id] = {
            'tag'   : tag,
            **run_params,
            # ---- Santa ----
            'ss_event'       : ev_id,
            'ss_break_ini'   : ev.get('break_ini'),
            'ss_break_end'   : ev.get('break_end'),
            'ss_par1_break'  : ev.get('Par1_fragments',''),
            'ss_par2_break'  : ev.get('Par2_fragments',''),
            'ss_det_P1'      : ev.get('det_par1',''),
            'ss_det_P2'      : ev.get('det_par2',''),
            'ss_par1_tags'   : ev.get('Par1_tags',''),
            'ss_par2_tags'   : ev.get('Par2_tags',''),
            'ss_par1_match'  : ";".join(ev.get('par1_seqs',[])),
            'ss_par2_match'  : ";".join(ev.get('par2_seqs',[])),
            'ss_recs'        : ";".join(ev.get('recombinants',[])),
        }

    rows = []                      # they will be accumulated here

    # ======== CFML ====================================================
    for ev_id, s_ev in santa.items():
        span = santa_span(s_ev)
        det_ini, det_end = s_ev.get('break_ini'), s_ev.get('break_end')

        ranked = []
        if span and cfml:
            for node, cd in cfml.items():
                # ─────——  PATCH "[]" → [node]  —────——
                # Some terminal nodes carry the leaf label
                # as 'event' and NOT an internal sequence list.
                if not cd.get('sequences'):
                    cd['sequences'] = [str(node)]
                # ── if the sequence list is empty,
                #    use the event id as unique marker ──
                if not cd.get('sequences'):
                    cd['sequences'] = [str(ev_id)]

                ov_nt  = interval_overlap(det_ini,det_end,
                                           cd['break_ini'],cd['break_end'])
                ov_pct = 100*ov_nt/span
                ranked.append((ov_pct,ov_nt,node,cd))
            ranked.sort(key=lambda x: x[0], reverse=True)

        if not ranked:
            ranked = [(0,0,'ND',{})]

        ov_pct, ov_nt, node, cd = ranked[0]
        rk = 1
        ext = {
            'tool'              : 'CFML',
            'rank'              : rk,
            'ext_id'            : node,
            'ext_break_ini'     : cd.get('break_ini',''),
            'ext_break_end'     : cd.get('break_end',''),
            'break_overlap_nt'  : ov_nt,
            'break_overlap_pct' : round(ov_pct,2),
            'rec_overlap'       : len(set(cd.get('sequences',[])) &
                                      set(s_ev.get('recombinants_num',[]))),
            }
        rows.append({**base_for_event[ev_id], **ext})

    # ======== Bacter ==================================================
    for ev_id, s_ev in santa.items():
        s_rec = set(s_ev.get('recombinants_num',[]))
        ranked = []
        if bacter:
            for lab, bd in bacter.items():
                all_rec = set(bd['group1']) | set(bd['group2'])
                ov = len(all_rec & s_rec)
                ranked.append((ov, lab, bd))
            ranked.sort(key=lambda x: x[0], reverse=True)

        if not ranked or ranked[0][0] == 0:
            ranked = [(0,'ND',{})]

        for rk,(ov,lab,bd) in enumerate(ranked,1):
            ext = {
                'tool'        : 'Bacter',
                'rank'        : rk,
                'ext_id'      : lab,
                'break_overlap_nt'  : '',
                'break_overlap_pct' : '',
                'rec_overlap' : ov,
            }
            rows.append({**base_for_event[ev_id], **ext})

    # ======== RDP =====================================================
    for ev_id, s_ev in santa.items():
        span = santa_span(s_ev)
        det_ini, det_end = s_ev.get('break_ini'), s_ev.get('break_end')

        ranked = []

        genome_len = max(evv.get('break_end', 0) for evv in santa.values())
        if span and rdp and 'NRD' not in rdp:
            for (b,e), g in rdp.items():
                ov_nt  = interval_overlap(det_ini,det_end,b,e)
                ov_pct = 100*ov_nt/span
                ranked.append((ov_pct,ov_nt,b,e,g))
            ranked.sort(key=lambda x: x[0], reverse=True)

        if not ranked:
            ranked = [(0,0,'ND','ND',{})]

        for rk,(ov_pct,ov_nt,b,e,g) in enumerate(ranked,1):
            p_santa = set(s_ev.get('par1_seqs',[])) | set(s_ev.get('par2_seqs',[]))
            ext = {
                'tool'              : 'RDP',
                'rank'              : rk,
                'ext_id'            : f"{b}-{e}" if b!='ND' else 'ND',
                'ext_break_ini'     : b,
                'ext_break_end'     : e,
                'break_overlap_nt'  : ov_nt,
                'break_overlap_pct' : round(ov_pct,2),
                'rec_overlap'       : len(set(g.get('recombinant',[])) &
                                          set(s_ev.get('recombinants',[]))),
                'par_minor_overlap' : len(set(g.get('minor',[])) & p_santa),
                'par_major_overlap' : len(set(g.get('major',[])) & p_santa),
            }
            rows.append({**base_for_event[ev_id], **ext})

    # orden final estable
    rows.sort(key=lambda r:(r['ss_event'], r['tool'], r['rank']))
    return rows


# ---------------------------------------------------------------------------
# ➋  "Raw" DUMP  – all events to a single Excel
# ---------------------------------------------------------------------------
def write_events_excel(tag: str,
                       santa: dict[int, dict],
                       cfml: dict[str, dict],
                       bacter: dict[str, dict],
                       rdp: dict | None = None,
                       run_params: dict | None = None,
                       path: Path = Path("results/recombination_summary.xlsx"),
                       sheet: str = "events_by_tool") -> None:
    """
    Dumps all events *as-is* in the `events_by_tool` sheet
    (if the sheet already exists, adds them below).
    """
    rows = []
    base = {"tag": tag, **(run_params or {})}   # ← tag + parameters
    for ev_id, d in (santa  or {}).items():
        rows.append({**base, "tool": "Santa", "event_id": ev_id, **d})
    for node, d in (cfml or {}).items():
        # ─── PATCH CFML: if the list is empty we use the id itself ───
        if not d.get("sequences"):
            d["sequences"] = [str(node)]
        rows.append({**base, "tool": "CFML", "event_id": node, **d})
    for lab  , d in (bacter or {}).items():
        rows.append({**base, "tool": "Bacter", "event_id": lab , **d})
    if rdp and 'NRD' not in rdp:
        for (b,e), d in rdp.items():
            rows.append({**base, "tool": "RDP",
                         "event_id": f"{b}-{e}", **d})

    if not rows:
        return None

    df = pd.DataFrame(rows)
    append_df_to_excel(df, path, sheet)
    return df

# ──────────────────────────────────────────────────────────────
#  NEW IMPLEMENTATION ─ parent_set()
#     · Merges Santa's P1/P2 parents
#     · If include_recombinants=True, adds recombinants
#       (useful for treating them as potential parentals).
# ──────────────────────────────────────────────────────────────
def parent_set(ev: dict, *, include_recombinants: bool = False) -> set[str]:
    pars = set(ev.get("par1_seqs", [])) | set(ev.get("par2_seqs", []))
    if include_recombinants:
        pars |= {str(x) for x in ev.get("recombinants_num", [])}
    return pars

# ---------------------------------------------------------------------------
# ➊  GROUP the three Santa classes -------------------------------
#     · P1  = par1_seqs
#     · P2  = par2_seqs
#     · REC = recombinants_num   (cast → str to cross with other lists)
# ---------------------------------------------------------------------------
def santa_groups(ev: dict) -> dict[str, set[str]]:
    return {
        "P1" : set(ev.get("par1_seqs",      [])),
        "P2" : set(ev.get("par2_seqs",      [])),
        "REC": {str(x) for x in ev.get("recombinants_num", [])},
    }


# ---------------------------------------------------------------------------
# ➋  GIVEN THE OVERLAP MATRIX Santa-vs-tool,
#     assigns each Santa group to the external group with greatest intersection
#     (breaks ties by alphabetical order of external group name)
# ---------------------------------------------------------------------------
def best_matches(santa_dict: dict[str, set[str]],
                 tool_dict : dict[str, set[str]],
                 nd_val: str = "ND") -> dict[str, str]:
    """
    Assigns each Santa group to the external group with greatest intersection,
    using each external group **at most once**.
    If there aren't enough external groups, leftover Santa groups are marked with `nd_val`.
    """
    remaining_t = set(tool_dict)              # "free" external groups
    remaining_s = set(santa_dict)             # P1, P2, REC to assign
    match      : dict[str,str] = {}

    # build all possible intersections
    candidates = []
    for s_name, s_set in santa_dict.items():
        for t_name, t_set in tool_dict.items():
            candidates.append((len(s_set & t_set), s_name, t_name))

    # greedy: largest intersection first
    for _, s_name, t_name in sorted(candidates, reverse=True):
        if s_name in remaining_s and t_name in remaining_t:
            match[s_name] = t_name
            remaining_s.remove(s_name)
            remaining_t.remove(t_name)

    # what remains unassigned → ND
    for s_name in remaining_s:
        match[s_name] = nd_val

    return match

# ---------------------------------------------------------------------------
# ➌  NEW write_comparisons_excel  (overwrites the previous one)
#     Columns:
#        tag, event, tool, ext_id,
#        par1_match_n, par1_match_pct,
#        par2_match_n, par2_match_pct,
#        rec_overlap, rec_mis_as_par, par_mis_as_rec,
#        break_overlap_nt, break_overlap_pct
# ---------------------------------------------------------------------------
def write_comparisons_excel(tag: str,
                            santa:  dict[int, dict],
                            cfml:   dict[str, dict],
                            bacter: dict[str, dict],
                            rdp:    dict | None = None,
                            run_params: dict | None = None,
                            *,
                            skip_caret: bool = False,
                            write_no_detection: bool = False,
                            only_rdp: bool = False,
                            path:   Path = Path("results/santa_vs_tools.xlsx"),
                            sheet:  str  = "santa_vs_tools") -> None:

    rows = []

    # genomic length (for Dice adj.)
    if santa:
        genome_len = max(ev.get('break_end', 0) or 0 for ev in santa.values())
    else:
        # If no santa data, try to get from RDP
        if rdp and 'NRD' not in rdp:
            genome_len = max(max(k) for k in rdp if isinstance(k, tuple)) if rdp else 0
        else:
            genome_len = 0

    # ─── If write_no_detection and no Santa, write only available tools ───
    if write_no_detection and not santa:
        base_info = {'tag': tag, **(run_params or {})}
        
        # Write only RDP events without comparisons
        if rdp and 'NRD' not in rdp:
            for rk, ((b, e), g) in enumerate(sorted(rdp.items()), 1):
                _clean_rdp_lists(g, skip_caret=True)
                rows.append({
                    **base_info, 'event': None, 'tool': 'RDP',
                    'ext_id': f"{b}-{e}",
                    'par1_match_n': '',
                    'par1_match_pct': '',
                    'par1_missing_santa': '',
                    'par1_missing_tool': '',
                    'par2_match_n': '',
                    'par2_match_pct': '',
                    'par2_missing_santa': '',
                    'par2_missing_tool': '',
                    'rec_match_n': '',
                    'rec_match_pct': '',
                    'rec_missing_santa': '',
                    'rec_missing_tool': '',
                    'break_overlap_nt': '',
                    'break_overlap_pct': '',
                    'break_bp_sizes': '',
                    'nonmatch_sim_prod_p1': '',
                    'nonmatch_sim_prod_p2': '',
                })
        
        # Write CFML events without comparisons (if only_rdp is not active)
        if cfml and not only_rdp:
            for rk, (node, cd) in enumerate(cfml.items(), 1):
                rows.append({
                    **base_info, 'event': None, 'tool': 'CFML',
                    'ext_id': node,
                    'par1_match_n': '',
                    'par1_match_pct': '',
                    'par1_missing_santa': '',
                    'par1_missing_tool': '',
                    'par2_match_n': '',
                    'par2_match_pct': '',
                    'par2_missing_santa': '',
                    'par2_missing_tool': '',
                    'rec_match_n': '',
                    'rec_match_pct': '',
                    'rec_missing_santa': '',
                    'rec_missing_tool': '',
                    'break_overlap_nt': '',
                    'break_overlap_pct': '',
                    'break_bp_sizes': '',
                    'nonmatch_sim_prod_p1': '',
                    'nonmatch_sim_prod_p2': '',
                })
        
        # Write Bacter events without comparisons (if only_rdp is not active)
        if bacter and not only_rdp:
            for rk, (lab, bd) in enumerate(bacter.items(), 1):
                rows.append({
                    **base_info, 'event': None, 'tool': 'Bacter',
                    'ext_id': lab,
                    'par1_match_n': '',
                    'par1_match_pct': '',
                    'par1_missing_santa': '',
                    'par1_missing_tool': '',
                    'par2_match_n': '',
                    'par2_match_pct': '',
                    'par2_missing_santa': '',
                    'par2_missing_tool': '',
                    'rec_match_n': '',
                    'rec_match_pct': '',
                    'rec_missing_santa': '',
                    'rec_missing_tool': '',
                    'break_overlap_nt': '',
                    'break_overlap_pct': '',
                    'break_bp_sizes': '',
                    'nonmatch_sim_prod_p1': '',
                    'nonmatch_sim_prod_p2': '',
                })
        
        # If there are rows, write and return
        if rows:
            new_df = pd.DataFrame(rows)
            path.parent.mkdir(parents=True, exist_ok=True)
            if path.exists() and sheet in pd.ExcelFile(path).sheet_names:
                prev_df = pd.read_excel(path, sheet_name=sheet)
                prev_df = prev_df[prev_df["tag"] != tag]
                out_df = pd.concat([prev_df, new_df], ignore_index=True)
            else:
                out_df = new_df
            
            if path.exists():
                with pd.ExcelWriter(path, engine="openpyxl", mode="a") as wr:
                    if sheet in wr.book.sheetnames:
                        del wr.book[sheet]
                    out_df.to_excel(wr, sheet_name=sheet, index=False)
            else:
                out_df.to_excel(path, sheet_name=sheet, index=False)
            
            print(green(f"✓ Sheet '{sheet}' updated (tag {tag}, no Santa) → {path}"))
            return
    
    # If no Santa and not write_no_detection, exit early
    # (we already handled the write_no_detection case above)

    # ───────── helpers for new metrics ──────────────────
    def _sim_score(ev: dict, sid: str) -> float:
        """
        Returns the similarity (0-1) of *sid* according to the
        Sim__seq_PAR1 / PAR2 fields of the Santa event; 0.0 if it doesn't appear.
        """
        for key in ("Sim__seq_PAR1", "Sim__seq_PAR2"):
            for item in ev.get(key, []):
                seq, pct = item.split(":", 1)
                if seq == sid:
                    return float(pct.rstrip("%"))/100.0
        return 0.0

    # --- more generic wrapper: we already use sim_prod() defined above ---
    def sim_product(sim_map, santa_set, tool_set):
        """
        Similarity product for sequences that **are in the
        tool but not in Santa**.

          · If there are *no* extras                 → 'ND'
          · If there are extras *but* we have NO stored similarity
            (sim_map empty or without the ids)      → 'ND'
          · Otherwise                                → ∏ sim(i)  (rounded)
        """
        extras = tool_set - santa_set
        if not extras:
            return "ND"

        # is there at least one known similarity?
        known = [sim_map.get(sid) for sid in extras if sid in sim_map]
        if not known:                          # all unknown
            return "ND"

        prod = 1.0
        for val in known:
            prod *= val
        return round(prod, 6)

    # ---------- loop per Santa event ---------------------------------
    base_info = {'tag': tag, **(run_params or {})}   # ← common to all rows
    for ev_id, s_ev in santa.items():
        S  = santa_groups(s_ev)
        #print(s_ev.keys())          # do 'det_P1' and 'det_P2' appear?
        #print(repr(s_ev.get('det_par1')))
        #print(repr(s_ev.get('det_par2')))
        # ---------- possible Santa intervals for breakpoints ----------
        #    We take ALL alternatives from Par1 *and* Par2
        raw_list = []
        for key in ('det_par1','det_par2'):
            #print(f"WE ARE HERE: {key}")
            fr = s_ev.get(key,'')
            if fr:
                raw_list.extend(fr.split('/'))

        s_int_alts : list[list[tuple[int,int]]] = []
        for alt in raw_list:
            ints = []
            for frag in alt.split(','):
                if frag and '-' in frag: # HERE NDs ARE REMOVED
                    a,b = map(int, frag.split('-'))
                    ints.append((a,b))
            if ints:
                s_int_alts.append(ints)
                #print(f"WE ARE HERE: {s_int_alts}")
        if not s_int_alts:                       # fallback: use break_ini/break_end
            s_int_alts = [[(s_ev.get('break_ini'), s_ev.get('break_end'))]]
        SIM_P1 = _parse_sim_list(s_ev.get('Sim__seq_PAR1', []))
        SIM_P2 = _parse_sim_list(s_ev.get('Sim__seq_PAR2', []))

        s_ini, s_end = s_ev.get('break_ini'), s_ev.get('break_end')
        span = (s_end - s_ini + 1) if s_ini and s_end else None

        # ----------------------------------------------------------------
        # 1)  CFML  (single group: cd['sequences'])
        # ----------------------------------------------------------------
        if cfml and not only_rdp:
            for rk, (node, cd) in enumerate(cfml.items(), 1):
                # ─── PATCH CFML: guarantees at least one identifier ───
                seqs_cfml = cd.get("sequences", [])
                if not seqs_cfml:
                    seqs_cfml = [str(node)]             # terminal node = id
                cd["sequences"] = seqs_cfml             # ← update the dict!

                T = {"NODE": set(map(str, seqs_cfml))}

                assign = best_matches(S, T)

                # ── intersection/difference metrics ───────────────────────────────
                t_p1 = T.get(assign['P1'], set())     # ← use set() if the group doesn't exist
                t_p2 = T.get(assign['P2'], set())
                t_rec = T.get(assign['REC'], set())

                p1_n, p1_miss_S, p1_miss_T, p1_pct       = _set_stats(S['P1'],  T.get(assign['P1'],  set()))
                p2_n, p2_miss_S, p2_miss_T, p2_pct       = _set_stats(S['P2'],  T.get(assign['P2'],  set()))
                rec_n, rec_miss_S, rec_miss_T, rec_pct   = _set_stats(S['REC'], T.get(assign['REC'], set()))


                # --- breakpoint overlap (best of all alternatives) --
                # ------------------- breakpoints de CFML --------------------
                #  *We prefer* those from det_par1/det_par2 (break_1/2);
                #  if they don't exist, fall back to the old break_ini / break_end.
                print(F"\n\nComparing detectable breakpoints of Santa: {s_int_alts} with those of the tool: {cd.get('break_ini')} – {cd.get('break_end')} (and their complements)")
                bp_ov_nt, bp_pct, lenS, lenT      = _best_break_overlap(
                    s_int_alts,
                    cd.get('break_ini'), cd.get('break_end'),
                    genome_len
                )



                nms_p1   = sim_product(SIM_P1, S['P1'], T.get(assign['P1'], set()))
                nms_p2   = sim_product(SIM_P2, S['P2'], T.get(assign['P2'], set()))
                nms_rec  = sim_product({},      S['REC'], T.get(assign['REC'], set()))


                rows.append({
                    **base_info, 'event': ev_id, 'tool': 'CFML',
                    'ext_id': node,
                    'par1_match_n'        : p1_n,
                    'par1_missing_santa'  : p1_miss_S,
                    'par1_missing_tool'   : p1_miss_T,
                    'par1_match_pct'      : p1_pct,
                    'par2_match_n'        : p2_n,
                    'par2_missing_santa'  : p2_miss_S,
                    'par2_missing_tool'   : p2_miss_T,
                    'par2_match_pct'      : p2_pct,
                    'rec_match_n'         : rec_n,
                    'rec_missing_santa'   : rec_miss_S,
                    'rec_missing_tool'    : rec_miss_T,
                    'rec_match_pct'       : rec_pct,
                    'nonmatch_sim_prod_p1':  nms_p1,
                    'nonmatch_sim_prod_p2':  nms_p2,
                    #'nonmatch_sim_prod_rec': nms_rec,
                    'break_overlap_nt'               : bp_ov_nt,
                    'break_overlap_pct'              : bp_pct,
                    'break_bp_sizes'                 : f"{lenS},{lenT}"

                })

        # ----------------------------------------------------------------
        # 2)  Bacter  (group1, group2, group3)
        # ----------------------------------------------------------------
        if bacter and not only_rdp:
            for rk, (lab, bd) in enumerate(bacter.items(), 1):
                T = {
                    'group1': set(bd.get('group1', [])),
                    'group2': set(bd.get('group2', [])),
                    'group3': set(bd.get('group3', [])),
                }

                assign = best_matches(S, T)

                # ── intersection/difference metrics ───────────────────────────────
                t_p1 = T.get(assign['P1'], set())     # ← use set() if the group doesn't exist
                t_p2 = T.get(assign['P2'], set())
                t_rec = T.get(assign['REC'], set())

                p1_n, p1_miss_S, p1_miss_T, p1_pct       = _set_stats(S['P1'],  T.get(assign['P1'],  set()))
                p2_n, p2_miss_S, p2_miss_T, p2_pct       = _set_stats(S['P2'],  T.get(assign['P2'],  set()))
                rec_n, rec_miss_S, rec_miss_T, rec_pct   = _set_stats(S['REC'], T.get(assign['REC'], set()))

                nms_p1   = sim_product(SIM_P1, S['P1'], T.get(assign['P1'], set()))
                nms_p2   = sim_product(SIM_P2, S['P2'], T.get(assign['P2'], set()))
                nms_rec  = sim_product({},      S['REC'], T.get(assign['REC'], set()))

                rows.append({
                    **base_info, 'event': ev_id, 'tool': 'Bacter',
                    'ext_id': lab,
                    'par1_match_n': p1_n,
                    'par1_match_pct': round(100*p1_n/len(S['P1']),2) if S['P1'] else 0,
                    'par1_missing_santa': p1_miss_S,
                    'par1_missing_tool' : p1_miss_T,
                    'par2_match_n': p2_n,
                    'par2_match_pct': round(100*p2_n/len(S['P2']),2) if S['P2'] else 0,
                    'par2_missing_santa': p2_miss_S,
                    'par2_missing_tool' : p2_miss_T,
                    'rec_match_n': rec_n,
                    'rec_match_pct': round(100*rec_n/len(S['REC']),2) if S['REC'] else 0,
                    'rec_missing_santa': rec_miss_S,
                    'rec_missing_tool' : rec_miss_T,
                    'nonmatch_sim_prod_p1':  nms_p1,
                    'nonmatch_sim_prod_p2':  nms_p2,
                    #'nonmatch_sim_prod_rec': nms_rec,
                    'break_overlap_nt': '',    # Bacter no da bp
                    'break_overlap_pct': ''
                })

        # ----------------------------------------------------------------
        # 3)  RDP  (minor, major, recombinant)
        # ----------------------------------------------------------------
        if rdp and 'NRD' not in rdp:
            for rk, ((b,e), g) in enumerate(sorted(rdp.items()), 1):
                major_ints = []
                if b > 1:
                    major_ints.append((1, b-1))
                if e < genome_len:
                    major_ints.append((e+1, genome_len))

                rdp_alt_ints = [[(b, e)], major_ints] if major_ints else [[(b, e)]]

                _clean_rdp_lists(g, skip_caret=True)

                T = {
                    'minor'      : clean_set(g.get('minor', []),      skip_caret),
                    'major'      : clean_set(g.get('major', []),      skip_caret),
                    'recombinant': clean_set(g.get('recombinant', []), skip_caret),
                }

                assign = best_matches(S, T)

                # ── intersection/difference metrics ───────────────────────────────
                t_p1 = T.get(assign['P1'], set())     # ← use set() if the group doesn't exist
                t_p2 = T.get(assign['P2'], set())
                t_rec = T.get(assign['REC'], set())

                p1_n, p1_miss_S, p1_miss_T, p1_pct       = _set_stats(S['P1'],  T.get(assign['P1'],  set()))
                p2_n, p2_miss_S, p2_miss_T, p2_pct       = _set_stats(S['P2'],  T.get(assign['P2'],  set()))
                rec_n, rec_miss_S, rec_miss_T, rec_pct   = _set_stats(S['REC'], T.get(assign['REC'], set()))

                # --- best overlap: ALWAYS evaluated, whether or not there is
                #     direct overlap between (b-e) and (s_ini-s_end) ----
                bp_ov_nt, bp_pct, bp_pct_norm = 0, 0.0, ''
                if span:
                    for alt in rdp_alt_ints:         # [(b,e)]  and  major_ints
                        for seg in alt:              # each individual segment
                            # RDP block: use seg[0], seg[1] (NOT cd - cd only exists in CFML block starting at line 2470)
                            print(F"\n\nComparing detectable breakpoints of Santa: {s_int_alts} with those of the tool {seg[0]} – {seg[1]} (and their complements)")
                            nt, pct, lenS_, lenT_   = _best_break_overlap(
                                s_int_alts, seg[0], seg[1], genome_len
                            )
                            if pct > bp_pct:         # save the *best*
                                bp_ov_nt, bp_pct, lenS, lenT = (
                                    nt, pct, lenS_, lenT_
                                )
                    if bp_pct:                       # normalized only if ≠0
                        bp_pct_norm = round(
                            bp_pct * min((s_end - s_ini + 1), (e - b + 1))
                            / max((s_end - s_ini + 1), (e - b + 1)), 2)
                    else:
                        bp_pct = bp_pct_norm = bp_adj = ''
                else:
                    bp_pct = bp_pct_norm = bp_adj = ''

                nms_p1   = sim_product(SIM_P1, S['P1'], T.get(assign['P1'], set()))
                nms_p2   = sim_product(SIM_P2, S['P2'], T.get(assign['P2'], set()))
                nms_rec  = sim_product({},      S['REC'], T.get(assign['REC'], set()))

                rows.append({
                    **base_info, 'event': ev_id, 'tool': 'RDP',
                    'ext_id': f"{b}-{e}",
                    'par1_match_n': p1_n,
                    'par1_match_pct': p1_pct,
                    'par1_missing_santa': p1_miss_S,
                    'par1_missing_tool' : p1_miss_T,
                    'par2_match_n': p2_n,
                    'par2_match_pct': p2_pct,
                    'par2_missing_santa': p2_miss_S,
                    'par2_missing_tool' : p2_miss_T,
                    'rec_match_n': rec_n,
                    'rec_match_pct': rec_pct,
                    'rec_missing_santa': rec_miss_S,
                    'rec_missing_tool' : rec_miss_T,
                    'break_overlap_nt': bp_ov_nt,
                    'break_overlap_pct':               bp_pct,
                    'break_bp_sizes'                 : f"{lenS},{lenT}",
                    'nonmatch_sim_prod_p1':  nms_p1,
                    'nonmatch_sim_prod_p2':  nms_p2,
                    #'nonmatch_sim_prod_rec': nms_rec
                })

    # ─── If write_no_detection and there's Santa but no tools, write empty rows ───
    if write_no_detection and santa and not rows:
        # No external tools, write Santa events with empty columns
        base_info = {'tag': tag, **(run_params or {})}
        for ev_id in santa.keys():
            rows.append({
                **base_info, 'event': ev_id, 'tool': 'ND',
                'ext_id': '',
                'par1_match_n': '',
                'par1_match_pct': '',
                'par1_missing_santa': '',
                'par1_missing_tool': '',
                'par2_match_n': '',
                'par2_match_pct': '',
                'par2_missing_santa': '',
                'par2_missing_tool': '',
                'rec_match_n': '',
                'rec_match_pct': '',
                'rec_missing_santa': '',
                'rec_missing_tool': '',
                'break_overlap_nt': '',
                'break_overlap_pct': '',
                'break_bp_sizes': '',
                'nonmatch_sim_prod_p1': '',
                'nonmatch_sim_prod_p2': '',
            })

    # ----------------------- dump to Excel ----------------------------
    if not rows:
        print("-- No comparisons to write.")
        return

    new_df = pd.DataFrame(rows)
    path.parent.mkdir(parents=True, exist_ok=True)

    # read previous content (if exists) and discard the old tag
    if path.exists() and sheet in pd.ExcelFile(path).sheet_names:
        prev_df = pd.read_excel(path, sheet_name=sheet)
        prev_df = prev_df[prev_df["tag"] != tag]
        out_df  = pd.concat([prev_df, new_df], ignore_index=True)
    else:
        out_df = new_df

    # ───── write the workbook ──────────────────────────────────
    if path.exists():
        # file already exists → append mode, remove old sheet
        with pd.ExcelWriter(path, engine="openpyxl", mode="a") as wr:
            if sheet in wr.book.sheetnames:
                del wr.book[sheet]          # delete old sheet
            out_df.to_excel(wr, sheet_name=sheet, index=False)
    else:
        # file does NOT exist → create from scratch
        out_df.to_excel(path, sheet_name=sheet, index=False)

    print(green(f"✓ Sheet '{sheet}' updated (tag {tag}) → {path}"))

# ---------------------------------------------------------------------------
# ➍  Global metrics per tool (breakpoints, parentals, recs)
# ---------------------------------------------------------------------------
def write_tool_metrics(df: pd.DataFrame,
                       csv_path: Path = Path("results/performance_summary.csv")):
    """
    Accepts the complete DataFrame returned by write_event_summary and
    calculates simple performance metrics per tool:
        • number of matched events
        • average % of breakpoint overlap
        • average number of shared recombinants
    """
    if df.empty:
        return

    agg = (df.groupby('tool')
             .agg(events          = ('ss_event','nunique'),
                  mean_bp_pct     = ('break_overlap_pct','mean'),
                  mean_rec_shared = ('rec_overlap','mean'))
             .reset_index())

    csv_path.parent.mkdir(exist_ok=True)
    mode = 'a' if csv_path.exists() else 'w'
    header = not csv_path.exists()
    agg.to_csv(csv_path, mode=mode, header=header, index=False)

# ─────────────────── Integrate RDP in Excel ──────────────────────────
#  ➟ re-reads dictionaries generated by simulation and
#    calls write_events_excel / write_comparisons_excel
def integrate_rdp_results(tag: str,
                          rdp_events: dict,
                          par_threshold: float = 1e-8,
                          only_rdp: bool = False,
                          write_no_detection: bool = False):
    run_dir = Path("results") / tag
    if not run_dir.exists():
        print(f"[WARN] Simulation directory for {tag} not found – skipped.")
        return
    
    if not only_rdp:
        # ─── Normal mode: extract all results ───
        santa, cfml, bacter = extract_recombination_results(
            run_dir, par_threshold=par_threshold)

        # ───── determine genomic length (max. break_end from Santa) ─────
        genome_len = max(ev.get('break_end', 0) or 0 for ev in santa.values())
        if not genome_len:                        # backup → RDP data
            if rdp_events and 'NRD' not in rdp_events:
                genome_len = max(max(k) for k in rdp_events if isinstance(k, tuple))
            else:
                genome_len = 0
        
        run_params = parse_cfg_params(run_dir / "santa" / "cfg.xml")
    else:
        # ─── only_rdp mode: extract only Santa (ignore cfml/bacter) ───
        santa = extract_santa_only(run_dir)
        cfml = {}
        bacter = {}
        
        # ───── determine genomic length (prefer Santa, then RDP) ─────
        if santa:
            genome_len = max(ev.get('break_end', 0) or 0 for ev in santa.values())
        elif rdp_events and 'NRD' not in rdp_events:
            genome_len = max(max(k) for k in rdp_events if isinstance(k, tuple))
        else:
            genome_len = 0
        
        # Try to get run_params, but don't fail if it doesn't exist
        cfg_xml = run_dir / "santa" / "cfg.xml"
        run_params = parse_cfg_params(cfg_xml) if cfg_xml.exists() else {}

    # ───── add bp_minor / bp_major to RDP dicts ──────────────────
    if rdp_events and 'NRD' not in rdp_events:
        for (b, e), d in rdp_events.items():
            d['bp_minor'] = f"{b}-{e}"
            left  = f"1-{b-1}"         if b > 1           else ''
            right = f"{e+1}-{genome_len}" if e < genome_len else ''
            d['bp_major'] = ",".join(x for x in (left, right) if x)

            _clean_rdp_lists(d, skip_caret=True)

    write_events_excel(tag, santa, cfml, bacter, rdp_events, run_params=run_params)
    write_comparisons_excel(tag, santa, cfml, bacter, rdp_events,
                            run_params=run_params, skip_caret=_SKIP_CARET_IN_COMP,
                            write_no_detection=write_no_detection,
                            only_rdp=only_rdp)

    print(green(f"✓ Integrated RDP for {tag} → Excel updated"))


def create_cluster_script(args):
    py_file = Path(sys.argv[0]).name          # this .py you're using
    mem_raw = args.cluster_mem.upper()        # e.g. 32G
    mem_num = "".join(filter(str.isdigit, mem_raw))  # 32 for --beast_mem

    script_text = f"""#!/bin/bash
#SBATCH --job-name={args.cluster_jobname}
#SBATCH --time={args.cluster_time}
#SBATCH --mem={mem_raw}
#SBATCH --cpus-per-task={args.cluster_cpus}
#SBATCH --array={args.cluster_array}
#SBATCH --output=logs/{args.cluster_jobname}_%A_%a.out
#SBATCH --error=logs/{args.cluster_jobname}_%A_%a.err

RUN_ID=$SLURM_ARRAY_TASK_ID

python3 {py_file} \\
       --start "$RUN_ID" --end "$RUN_ID" \\
       --threads {args.cluster_cpus} \\
       --beast_mem {mem_num}
"""

    cwd = Path.cwd()
    (cwd / "logs").mkdir(exist_ok=True)
    out_path = cwd / "submit_cluster.sh"
    out_path.write_text(script_text)
    out_path.chmod(0o755)

    print("The 'logs' directory has been created; it stores console output "
          "from cluster nodes.")
    print(f"'submit_cluster.sh' generated in {cwd}. Submit it with: sbatch submit_cluster.sh")
    sys.exit(0)

def write_event_summary(tag: str,
                        run_params: dict,
                        santa: dict[int, dict],
                        cfml: dict[str, dict],
                        bacter: dict[str, dict],
                        rdp: dict[tuple[int, int], dict] | None = None):

    """
    Merges results dictionaries, builds the rows
    and writes/adds them to the master CSV. Returns the newly
    generated DataFrame (in case you want it in memory).
    """
    rows = build_event_rows_v3(tag, run_params, santa, cfml, bacter, rdp)
    return _pd.DataFrame(rows)

# ─────────────────────────────── main ──────────────────────────────────
def main():
    p = argparse.ArgumentParser(description="Run full simulation and recombination analysis pipeline using Santa-Sim, IQ-TREE, ClonalFrameML, BEAST-Bacter, and RDP.")

    # ───────────── General Execution Range ─────────────
    p.add_argument("--download-software", action="store_true",
                   help="Download and set up all required software in ./software/.")
    p.add_argument("--start", type=int, default=1,
                   help="Start index for simulation runs (default: 1).")
    p.add_argument("--end", type=int, default=1,
                   help="End index for simulation runs (default: 1).")
    p.add_argument("--num_seqs", type=int, default=50,
                   help="Number of sequences sampled per simulation (default: 50).")
    p.add_argument("--ref", default="NC_045512.2.fna",
                   help="Reference genome in FASTA format to append as 'root' (default: NC_045512.2.fna).")

    # ───────────── Santa-Sim Simulation Parameters ─────────────
    p.add_argument("--popSize", type=int, default=10000,
                   help="Santa-Sim population size (default: 10000).")
    p.add_argument("--mutationRate", default="1.1E-6",
                   help="Per-site mutation rate in Santa-Sim (default: 1.1E-6).")
    p.add_argument("--recombProb", type=float, default=0.002,
                   help="Genome-wide recombination probability (default: 0.002).")
    p.add_argument("--dualInfectProb", type=float, default=0.02,
                   help="Probability of dual infection (default: 0.02).")
    p.add_argument("--genCount", type=int, default=10000,
                   help="Number of generations to simulate (default: 10000).")

    # ───────────── Software Paths ─────────────
    p.add_argument("--soft_dir", default="software",
                   help="Base directory for all installed software (default: software/).")
    p.add_argument("--santa_bin", default="software/santa-sim-Recomb_and_align_Luis_MOD/dist/santa.jar",
                   help="Path to Santa-Sim .jar executable.")
    p.add_argument("--iqtree_bin", default="software/iqtree-2.2.0-Linux/bin/iqtree2",
                   help="Path to IQ-TREE binary.")
    p.add_argument("--cfml_bin", default="software/ClonalFrameML/src/ClonalFrameML",
                   help="Path to ClonalFrameML binary.")
    p.add_argument("--bacter_bin", default="software/beast2.7.7/bin/beast",
                   help="Path to BEAST binary (with Bacter plugin).")
    p.add_argument("--condensetree_bin", default="software/beast2.7.7/bin/applauncher",
                   help="Path to BEAST app launcher (used for tree summarization).")

    # ───────────── BEAST-Bacter Parameters ─────────────
    p.add_argument("--bacter_chain", type=int, default=50000,
                   help="Total MCMC chain length for BEAST-Bacter (default: 50000).")
    p.add_argument("--bacter_delta", type=float, default=100.0,
                   help="Initial delta value (mean length of recombination fragments, in bp) (default: 100.0).")
    p.add_argument("--bacter_rho", type=float, default=0.002,
                   help="Initial rho value (recombination rate parameter) (default: 0.002).")
    p.add_argument("--beast_mem", type=int, default=3.5,
                   help="Maximum heap size for BEAST JVM in GB (default: 3.5).")

    # ───────────── ESS Calculation Parameters ─────────────
    p.add_argument("--burnin", type=float, default=0.1,
                   help="Burn-in fraction (0–1) or number of samples to discard for ESS computation (default: 0.1).")

    # ───────────── Result Extraction Modes ─────────────
    p.add_argument("--extract-results", metavar="DIR",
                   help="Extract recombination events from a completed run directory.")
    p.add_argument("--rdp-dir", "--rdp", metavar="DIR",
                   help="Parse all *.fasta.csv recombination results from RDP in the given directory.")
    p.add_argument("--only-RDP", action="store_true",
                   help="When processing RDP results, extract Santa and RDP events only, ignoring cfml and bacter directories.")
    p.add_argument("--write-also-no-detection", action="store_true",
                   help="Write santa_vs_tools.xlsx even when some tools have no detections, leaving columns empty for missing tools.")

    # ───────────── Parallelization / Resources ─────────────
    p.add_argument("--threads", type=int, default=1,
                   help="Number of threads for IQ-TREE, ClonalFrameML, and BEAST (default: 1).")

    # ───────────── SLURM Cluster Script Generation ─────────────
    p.add_argument("--script_cluster", action="store_true",
                   help="Generate a SLURM sbatch script (submit_cluster.sh) and exit.")
    p.add_argument("--cluster_jobname", default="RDP_job",
                   help="SLURM job name (--job-name) (default: RDP_job).")
    p.add_argument("--cluster_time", default="6:00:00",
                   help="SLURM time limit in HH:MM:SS format (default: 4:00:00).")
    p.add_argument("--cluster_mem", default="32G",
                   help="Memory allocation per SLURM task (e.g., 32G).")
    p.add_argument("--cluster_cpus", type=int, default=16,
                   help="Number of CPUs per SLURM task (--cpus-per-task) (default: 16).")
    p.add_argument("--cluster_array", default="1-100",
                   help="SLURM array range for parallel runs (default: 1-100).")

    # ───────────── Execution Mode Flags ─────────────
    p.add_argument("--resume-beast-only", action="store_true",
                   help="Skip SantaSim, IQ-TREE, and ClonalFrameML. Only resume BEAST-Bacter and post-processing.")

    # ───────────── Matching / Filtering ─────────────
    p.add_argument("--par-threshold", type=float, default=1e-8,
                   help="Maximum p-distance allowed to consider a sequence as a matching parental (default: 1e-8).")
    p.add_argument("--exclude-RDP", default="",
                   help="Comma-separated list of RDP breakpoint flags (~,*,$,^) to ignore during parsing.")

    args = p.parse_args()

    # ─── breakpoint-flag handling (applies to parsing *and* comparisons) ───
    excluded_flags = {c.strip() for c in args.exclude_RDP.split(',') if c.strip()}
    global _SKIP_CARET_IN_COMP
    _SKIP_CARET_IN_COMP = '^' in excluded_flags

    # ─── standalone «--extract-results DIR»  ───────────────────────────────
    if args.extract_results:
        run_dir = Path(args.extract_results).resolve()
        tag     = run_dir.name

        santa_dict, cfml_dict, bacter_dict = extract_recombination_results(
            run_dir, par_threshold=args.par_threshold)

        # effective parameters saved in cfg.xml (if exists)
        cfg_xml    = run_dir / "santa" / "cfg.xml"
        run_params = parse_cfg_params(cfg_xml) if cfg_xml.exists() else {}

        # SAME output as in «resume-beast-only» mode
        write_events_excel(tag, santa_dict, cfml_dict, bacter_dict,
                           run_params=run_params)
        write_comparisons_excel(tag, santa_dict, cfml_dict, bacter_dict,
                                run_params=run_params,
                                skip_caret=_SKIP_CARET_IN_COMP)

        print(green(f"✓ Results for run '{tag}' extracted and written to Excel"))
        sys.exit(0)

    soft_dir = Path(args.soft_dir).resolve()

    if args.rdp_dir:
        parse_rdp_dir(Path(args.rdp_dir),
                      excluded_flags,
                      start_idx=args.start,
                      end_idx=args.end,
                      only_rdp=args.only_RDP,
                      write_no_detection=args.write_also_no_detection)
        sys.exit(0)

    # ───── software download ─────
    if args.download_software:
        download_software(soft_dir)
    
    # ───── cluster script creation ─────
    if args.script_cluster:
        create_cluster_script(args)

    # ───── executable paths ─────
    santa_jar  = Path(args.santa_bin).resolve()
    iqtree_bin = Path(args.iqtree_bin).resolve()
    cfml_bin   = Path(args.cfml_bin).resolve()
    bacter_bin = Path(args.bacter_bin).resolve()
    condensetree_bin = Path(args.condensetree_bin).resolve()

    root = Path("results");  root.mkdir(exist_ok=True)
    rdpdir = root / "RDP";    rdpdir.mkdir(exist_ok=True)

    ref_name = Path(args.ref).stem

    for idx in range(args.start, args.end + 1):
        tag       = f"{idx}_{args.num_seqs}_{ref_name}"
        run_dir   = root / tag
        santa_dir = run_dir / "santa"
        iq_dir    = run_dir / "iqtree"
        cfml_dir  = run_dir / "cfml"
        bacter_dir= run_dir / "bacter"

        # ────────────────────────────────────────────────────────────────
        #  BRANCH  A  →  resume only BEAST-Bacter (if --resume-beast-only)
        # ────────────────────────────────────────────────────────────────
        if args.resume_beast_only:
            bacter_xml = bacter_dir / f"{tag}.xml"
            if not bacter_xml.exists():
                sys.exit(f"{bacter_xml} not found; cannot resume.")

            # ---------- helper: check ESS already calculated -------------
            log_file = bacter_dir / "bacter.log"

            def ess_ok() -> tuple[bool, tuple[Optional[float], Optional[float]]]:
                """
                True if bacter.log exists and ESS(rho & delta) ≥ 200.
                Returns (flag_ok, (ESS_rho, ESS_delta)).
                """
                if not log_file.exists():
                    return False, (None, None)
                df  = read_beast_log(log_file)
                ess = compute_ess(df, args.burnin)
                r, d = ess.get("rho"), ess.get("delta")
                if r is None or d is None:
                    return False, (r, d)
                return (r >= 200 and d >= 200), (r, d)

            ok, (r_ess, d_ess) = ess_ok()
            if ok:
                print(f"✔ Enough ESS (rho={r_ess:.1f}, delta={d_ess:.1f}); "
                      "BEAST is not relaunched")
            else:
                print(f"▶  Resuming BEAST for run {tag}")

                # --- helper to launch BEAST with error control ----
                def run_beast(cmd):
                    rc = shell(cmd, cwd=bacter_dir, ignore_err=True)
                    if rc not in (0, 1, 137):
                        sys.exit(f"BEAST terminated with unexpected code {rc}")
                    return rc

                # first resumption
                run_beast([str(bacter_bin), "-threads", str(args.threads),
                           "-seed", str(idx), "-resume", str(bacter_xml.resolve())])

                # ─── ESS loop / resume ──────────────────────────────
                while True:
                    ok, (r_ess, d_ess) = ess_ok()
                    if ok:
                        print(f"Convergence achieved "
                              f"(rho = {r_ess:.1f}, delta = {d_ess:.1f})")
                        break
                    print(f"↻ Insufficient ESS (rho={r_ess}, delta={d_ess}); "
                          "resuming...", flush=True)
                    run_beast([str(bacter_bin), "-threads", str(args.threads),
                               "-seed", str(idx), "-resume",
                               str(bacter_xml.resolve())])

            # condense tree and extract events
            shell([str(condensetree_bin), "ACGAnnotator",
                   str((bacter_dir / "bacter.trees").resolve())], cwd=bacter_dir)

            santa_dict, cfml_dict, bacter_dict = extract_recombination_results(
                    run_dir, par_threshold=args.par_threshold)
            
            # Re-read effective parameters saved in cfg.xml
            run_params = parse_cfg_params(santa_dir / "cfg.xml")

            # ────── NEW: unified excels ──────
            run_params = parse_cfg_params(santa_dir / "cfg.xml")          #  you already have it calculated before
            write_events_excel(tag, santa_dict, cfml_dict, bacter_dict, run_params=run_params)
            write_comparisons_excel(tag, santa_dict, cfml_dict, bacter_dict,
                                    run_params=run_params, skip_caret=_SKIP_CARET_IN_COMP)


        # (we NO longer generate recombination_events.csv)
            continue   # go to next idx  (END of branch A!!)
        # ────────────────────────────────────────────────────────────────
        #  BRANCH  B  →  full pipeline (what you already had)
        # ────────────────────────────────────────────────────────────────

        # create subdirectories if they don't exist
        if args.only_RDP:
            # Only create santa directory
            santa_dir.mkdir(parents=True, exist_ok=True)
        else:
            for d in (santa_dir, iq_dir, cfml_dir, bacter_dir):
                d.mkdir(parents=True, exist_ok=True)

        # ---------- 1) Santa‑Sim, 2) IQ‑TREE, 3) CFML, 4) BEAST ----------

        # 1) prepare XML inside each santa/ folder
        cfg_xml = santa_dir / "cfg.xml"
        create_santa_config_xml(
            args.ref, cfg_xml,
            args.popSize, args.mutationRate,
            args.recombProb, args.dualInfectProb,
            args.genCount, args.num_seqs
        )

        # convert to absolute:
        xml_abs = cfg_xml.resolve()

        # 2) execute Santa‑Sim
        shell([
            "java", "-jar", str(santa_jar),
            f"-seed={idx}",
            str(xml_abs)
        ], cwd=santa_dir)

        fasta  = santa_dir / "final_sample.fasta"

        withroot_fa = santa_dir / "final_sample_withroot.fasta"
        shutil.copy2(fasta, withroot_fa)

        ref_id = None
        with open(args.ref) as ref_in, open(withroot_fa, "a") as out:
            head = ref_in.readline().strip()
            if head.startswith(">"):
                ref_id = "root"           # ID sin ">"
                out.write(">root" + "\n")
            else:
                sys.exit("The --ref file is not FASTA")
            for ln in ref_in:
                out.write(ln)

        # S
        breakpoints_santa = santa_dir / "recombination_events_final_sample.txt"
        gen_seqs_santa    = santa_dir / "sequence_events_map_final_sample.txt"
        mlpref = iq_dir / tag

        # copy fasta to RDP6
        shutil.copy2(fasta, rdpdir / f"{tag}.fasta")

        if args.only_RDP:
            # ─── only_RDP mode: skip IQ-TREE, CFML and Bacter ───
            print(f"[INFO] --only-RDP active: skipping IQ-TREE, ClonalFrameML, and Bacter")
            santa_dict = extract_santa_only(run_dir)
            cfml_dict = {}
            bacter_dict = {}
            run_params = parse_cfg_params(cfg_xml)
        else:
            # ───── 2 · IQ‑TREE ─────
            shell([str(iqtree_bin), "-T", str(args.threads), "-s", str(withroot_fa), "-m", "TEST", "-nt", "AUTO", "-o", ref_id, "--seed", str(idx),"-pre", str(mlpref)])

            rooted_tree   = mlpref.with_name(mlpref.name + ".treefile")
            unroot_tree = mlpref.with_name(mlpref.name + "_unrooted.tree")
            t = Phylo.read(rooted_tree, "newick")
            t.rooted = False
            t.prune("root")
            Phylo.write(t, unroot_tree, "newick")

            # ───── 3 · ClonalFrameML ─────
            cfml_prefix = cfml_dir / tag
            shell([str(cfml_bin), str(unroot_tree), str(fasta), str(cfml_prefix)]) # SEED COMMAND DOES NOT WORK. THEORETICALLY ACCORDING TO THE TUTORIAL IT IS -S #N BUT IT FAILS BECAUSE IT DOESN'T RECOGNIZE THE ARGUMENT.

            # ───── 4 · Bacter ────────────────────────────────────────────────────
            bacter_xml = bacter_dir / f"{tag}.xml"
            build_bacter_xml(fasta, bacter_xml, tag,
                             args.bacter_chain, args.bacter_delta, args.bacter_rho)

            env_beast = os.environ.copy()
            env_beast["BEAST_OPTS"] = f"-Xmx{args.beast_mem}g -Xss256m"

            def run_beast(cmd):
                """
                Launches BEAST inside bacter_dir.
                Returns the exit-code but never raises exception for 1 or 137.
                """
                rc = shell(cmd, cwd=bacter_dir, env=env_beast, ignore_err=True)
                if rc not in (0, 1, 137):
                    sys.exit(f"BEAST terminated with unexpected code {rc}")
                return rc

            # 1st execution
            run_beast([str(bacter_bin), "-threads", str(args.threads),"-seed", str(idx), str(bacter_xml.resolve())])

            log_file = bacter_dir / "bacter.log"

            # ─── ESS loop / resume ────────────────────────────────────────────
            r = d = float("nan")
            while True:
                if not log_file.exists():
                    print("   (bacter.log not showing up yet, waiting…)")
                else:
                    df  = read_beast_log(log_file)
                    ess = compute_ess(df, args.burnin)
                    r, d = ess.get("rho"), ess.get("delta")
                    if r is not None and d is not None:
                        print(f"   ➤ ESS rho.t={r:.1f}, delta.t={d:.1f}")
                        if r >= 200 and d >= 200:
                            print(f"Convergence has been achieved with Rho = {r} and Delta = {d}", flush=True)
                            break
                print(f"   ↻ Insufficient ESS (rho = {r}; delta = {d}) or BEAST interrupted; resuming…", flush=True)
                run_beast([str(bacter_bin), "-threads", str(args.threads),"-seed", str(idx), "-resume", str(bacter_xml.resolve())])

            bater_trees_file = bacter_dir / "bacter.trees"
            condense_tree = bacter_dir / "summary.tree"

            shell([ str(condensetree_bin), "ACGAnnotator", "bacter.trees" ], cwd=bacter_dir)
            run_params = parse_cfg_params(cfg_xml)
            # 1) extract events from the just-finished replicate
            santa_dict, cfml_dict, bacter_dict = extract_recombination_results(run_dir, par_threshold=args.par_threshold)


        # ---------------- Save summaries and metrics ----------------
        if not args.only_RDP:
            df_summary = write_event_summary(tag, run_params,
                                             santa_dict, cfml_dict,
                                             bacter_dict, rdp=None)
            write_tool_metrics(df_summary,
                               Path("results/performance_summary.csv"))
        
        # Write Excels with the only_rdp flag if applicable
        write_events_excel(tag, santa_dict, cfml_dict, bacter_dict, run_params=run_params)
        write_comparisons_excel(tag, santa_dict, cfml_dict, bacter_dict,
                                run_params=run_params, skip_caret=_SKIP_CARET_IN_COMP,
                                write_no_detection=args.write_also_no_detection,
                                only_rdp=args.only_RDP)
    #     # 2) save/add to common Excel
    #     params = {
    #         'popSize'        : args.popSize,
    #         'mutationRate'   : args.mutationRate,
    #         'recombProb'     : args.recombProb,
    #         'dualInfectProb' : args.dualInfectProb,
    #         'genCount'       : args.genCount,
    #         'sampleSize'     : args.num_seqs,
    #     }
    #     rows = build_event_rows(tag, params,
    #                             santa_dict, cfml_dict, bacter_dict,
    #                             rdp=None,               # ← you don't have RDP here yet
    #                             par_threshold=args.par_threshold)
    #     append_event_rows(rows)     # ← writes to results/recombination_events.csv
    print_column_legend()
    print("\n\n✓ Pipeline completed")

# ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
